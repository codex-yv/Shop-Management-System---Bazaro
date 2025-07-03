import os 
import re
import time
import openpyxl
import string
from datetime import datetime, timedelta
from calendar import month_abbr
import sqlite3
import socket
import pymongo
import os
import pythoncom
import win32com.client as win32
import threading
from tkinter import*
from tkinter import ttk
import customtkinter as ctk
from tkinter import messagebox, filedialog, simpledialog
from tkcalendar import Calendar
from PIL import Image, ImageTk
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import random
import smtplib
from email.message import EmailMessage
from pathlib import Path
from dotenv import load_dotenv
import pandas as pd
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
import webbrowser
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk


error = 0
otpl =[]
new_email_list = []

func_list = ['dashboard']

inventory_func_list = ['None']

product_id_found = {
    'Update':False  ,
    'UID':000
}

choice_list = []

username_list = []

def check_internet(host="8.8.8.8", port=53, timeout=3):
    try:
        # Attempt to connect to a well-known DNS server (Google)
        socket.setdefaulttimeout(timeout)
        with socket.create_connection((host, port)):
            return True
    except OSError:
        return False

def update_day_date_time():
    now = datetime.now()
    
    day = now.strftime("%A") 
    date_time = now.strftime("%d/%m/%Y %H:%M:%S")
    day_label.config(text=day)
    date_time_label.config(text=date_time)

    win.after(1000, update_day_date_time)


def get_percentage_of_day():
    now = datetime.now()
    seconds_passed = now.hour * 3600 + now.minute * 60 + now.second
    total_seconds_in_day = 24 * 3600
    # print(now.day())
    return (seconds_passed / total_seconds_in_day) * 100

def update_progress():
    
    percentage = get_percentage_of_day()

    update_earnings()

    daily_ern = earnings.find_one({"Earning_ID":101}, {"Daily_Income":1})
    
    daily_income_profit.configure(text = daily_ern["Daily_Income"])

    # progress_var.set(percentage)
    daily_progress.set(get_percentage_of_day() / 100) 
    daily_progress_label.configure(text=f"{percentage:.2f}% day completed")
    win.after(3000, update_progress)

def generate_otp():
    return random.randint(100000, 999999)

week = {
    'Monday':1,
    'Tuesday':2,
    'Wednesday':3,
    'Thursday':4,
    'Friday':5,
    'Saturday':6,
    'Sunday':7
}

def get_fraction_of_week():
    global week
    
    now = datetime.now()
    day = now.strftime("%A") 
    
    day_count = week[day]
    
    weekly_progress_label.config(text = day)

    return (day_count/7)*100

def update_week_progress():
    
    day_frac = get_fraction_of_week()    
    
    weekly_ern = earnings.find_one({"Earning_ID":101}, {"Weekly_Income":1})
    
    weekly_income_profit.configure(text = weekly_ern["Weekly_Income"])
    
    weekly_progress.set(day_frac/100)
    
    win.after(3000, update_week_progress)

month_days = {
    1: 31,   # January
    2: 28,   # February (non-leap year)
    3: 31,   # March
    4: 30,   # April
    5: 31,   # May
    6: 30,   # June
    7: 31,   # July
    8: 31,   # August
    9: 30,   # September
    10: 31,  # October
    11: 30,  # November
    12: 31   # December
}
    
    
def get_monthly_percent():
    global month_days
    date = datetime.today().day

    month_num =  datetime.now().month
    days = month_days[month_num]
    monthly_progress_label.config (text=f"Day {date} of {days}")
    
    return (date/days)*100

def update_month_progress():
    
    month_frac = get_monthly_percent()
    
    monthly_ern = earnings.find_one({"Earning_ID":101}, {"Monthly_Income":1})
    
    monthly_income_profit.configure(text = monthly_ern["Monthly_Income"])
    
    monthly_progress.set(month_frac/100)

    win.after(3000, update_month_progress)
    
def update_server(dictonary):
    global client_info
    
    client_info.insert_one(dictonary)
    messagebox.showinfo('Verification Done', 'Sign Up Successful!')    
    
    
def check_server(chk_value, updt_value):
    global client_info
    
    find_val = client_info.find({}, {'name':chk_value}, )
    

def send_otp(email):
    global current_dir
    if check_internet():
        otp = generate_otp()
        otpl.insert(0,otp)
        
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        
        envars = current_dir/ ".env" # create .env file in current folder

        load_dotenv(envars)

        sender_email = os.getenv('EMAIL') #add your email in .env file
        sender_key = os.getenv('KEY') # add gamil app passwords in .env file

        server.login(sender_email, sender_key)
        to_mail = email

        msg = EmailMessage()

        msg['Subject'] = 'OTP verification'
        msg['From'] = 'ytgamings802212@gmail.com'
        msg['to'] = to_mail

        msg.set_content("Your OTP is " + str(otp))
        server.send_message(msg)
        messagebox.showinfo('OTP send', 'OTP send!')
    else:
        messagebox.showerror('Connection Error', 'Please Check your internet connection!')

def new_password():
    global client_info, new_email_list
    if len(username.get().strip())<4:
        messagebox.showerror('Reset Password', f"No username found with '{username.get()}'")
    else:
        find_val_username = client_info.find_one(
                {"username":username.get().strip() })
        if find_val_username:
            contentframe.pack_forget()
            pass_reset_frame.pack()
            find_email = client_info.find_one({'username':username.get().strip()}, {'email':1})
            new_email = find_email['email']
            new_email_list.insert(0, new_email)
            
            threading.Thread(target=send_otp, args=(new_email,), daemon=True).start()
            reset_canvas.create_text(1000, 350, text=f'OTP sent on {new_email}', font=('Poppins', 20), fill='#6F6F6F' )
            # print(email)
        else:
            messagebox.showerror('Reset Password', f"No username found with '{username.get()}'")

def sign_up():
    contentframe.pack_forget()
    sign_up_frame.pack()
    # sign_up_canvas.pack()

def try_login ():
    global username_list
    if check_internet():
        
        username_value = username.get().strip()
        password_value = password.get()
        
        client_info_dict = {
                'username':username_value,
                'password':password_value,
            }
            
        find_val_username = client_info.find_one(
                {"username": client_info_dict["username"]})
        
        if find_val_username:
            find_val_password = client_info.find_one(
                {"password": client_info_dict["password"]})
            if find_val_password:
                contentframe.pack_forget()
                dashboard_frame.pack()
                username_label.config(text= username_value)
                username_list.insert(0, username_value)
                update_day_date_time()
                update_progress()
                update_week_progress()
                update_month_progress()
                print(username_value,'\n',password_value)
            else:
                messagebox.showerror('Login Error', 'passowrd is wrong!')
        else:
            messagebox.showerror('Login Error', 'username is wrong!')
    else:
        messagebox.showerror('Connection Error', 'Please Check your internet connection!')

def reset_verify_otp():
    global otpl
    reset_otp_val = resetEmailotp.get()

    if str(otpl[0]) == reset_otp_val:
        pass_reset_frame.pack_forget()
        new_pass_frame.pack()
    else:
        messagebox.showerror('Verification Error', 'incorrect OTP, please resend the otp and enter the correct one!')

def reset_password():
    global client_info
    
    new_pass_update = client_info.update_one({'username': username.get().strip()}, {"$set":{'password':resetpass.get()}})

    if new_pass_update.matched_count>0:
        messagebox.showinfo('Password Reset', f'Your new password is {resetpass.get()}')
        new_pass_frame.pack_forget()
        contentframe.pack()

def reset_resend_otp():
    global new_email_list
    threading.Thread(target=send_otp, args=(new_email_list[0],), daemon=True).start()

def try_signup():
    global error
    username_value_signup =  username_signup.get().strip()
    email_value_signup = email_signup.get()
    phone_value_signup = phone_signup.get()
    password_value_signup = password_signup.get()
    
    if len(username_value_signup)<=2:
        messagebox.showerror('Incorrect Name', "Please enter your correct name!")
        error = error+1
    elif len(password_value_signup)<=4:
        messagebox.showwarning('Security Alert', 'Password length should atlest be 4')
        error = error+1
    else:
        pass
    
    
    email_condition = "^[a-z]+[\._]?[a-z 0-9]+[@]\w+[.]\w{2,3}$"

    if re.search(email_condition, email_value_signup):
        pass
    else:
        error = error+1
        messagebox.showerror('Wrong Email Format', f"'{email_value_signup}' has a wrong email format, Please check again!")

    mobile_condition = "^\d{10}$"

    if re.search(mobile_condition, phone_value_signup):
        pass
    else:
        error = error+1
        messagebox.showerror('Wrong Phone Format', f"'{phone_value_signup}' is wrong phone format!")

    if error == 0:
        sign_up_frame.pack_forget()
        verification_frame.pack()
        threading.Thread(target=send_otp, args=(email_value_signup,), daemon=True).start()
    else:
        error = 0
        messagebox.showerror('Recheck', 'Please recheck your credentials!')
            
    
def verify_otp():
    global otpl, client_info
    
    if email_otp.get() == str(otpl[0]):
        
        client_info_dict = {
            'username':username_signup.get().strip(),
            'password':password_signup.get(),
            'email':email_signup.get().strip(),
            'phone':int(phone_signup.get().strip())
        }
        
        find_val = client_info.find_one({
            "$or": [
                {"username": client_info_dict["username"]},
                {"phone": client_info_dict["phone"]},
                {"email": client_info_dict["email"]}
            ]
        })
        
        if find_val:
            messagebox.showerror('', 'username/phone/email already exist. Please Retry!')
        else:
            update_server(client_info_dict)
        
        
    else:
        messagebox.showerror('Verification Error', 'incorrect OTP, please resend the otp and enter the correct one!')


def resend_otp():
    threading.Thread(target=send_otp, args=(email_signup.get(),), daemon=True).start()
        
def get_back():
    verification_frame.pack_forget()
    sign_up_frame.pack()

def func_finder(prev_func_name):
    global inventory_func_list
    if prev_func_name == 'dashboard':
        dashboard_display.pack_forget()
    elif prev_func_name == 'inventory':
        inventory_display.pack_forget()
        if inventory_func_list[0] == 'add_stock':
            add_stock_frame.pack_forget()
        elif inventory_func_list[0] == 'update_stock':
            update_stock_frame.pack_forget()
        elif inventory_func_list[0] == 'analyse_stock':
            stock_analytics_frame.pack_forget()
        
    elif prev_func_name == 'alert':
        alert_display.pack_forget()
    elif prev_func_name == 'billing':
        billing_display.pack_forget()
    elif prev_func_name == 'supplier':
        supply_display.pack_forget()
    elif prev_func_name == 'history':
        history_display.pack_forget()
    elif prev_func_name == 'setting':
        setting_display.pack_forget()
    elif prev_func_name == 'cc':
        cc_display.pack_forget()
    
    
def dashboardFunction():
    global func_list
    prev_func_name = func_list[0]
    current_func_name = "dashboard"
    if prev_func_name == current_func_name:
        pass
    else:
        func_finder(prev_func_name)
        func_list.insert(0, current_func_name)
        dashboard_display.pack(side='left')


def inventoryFunction():
    global func_list
    prev_func_name = func_list[0]
    current_func_name = "inventory"
    if prev_func_name == current_func_name:
        pass
    else:
        func_finder(prev_func_name)
        func_list.insert(0, current_func_name)
        inventory_display.pack(side='left')

def alertFunction():
    global func_list
    prev_func_name = func_list[0]
    current_func_name = "alert"
    if prev_func_name == current_func_name:
        pass
    else:
        func_finder(prev_func_name)
        func_list.insert(0, current_func_name)
        alert_display.pack(side='left')
        insert_data_to_alert_treeview()
        insert_data_to_alert_treeview_stock()
        

def billingFunction():
    global func_list
    prev_func_name = func_list[0]
    current_func_name = "billing"
    if prev_func_name == current_func_name:
        pass
    else:
        func_finder(prev_func_name)
        func_list.insert(0, current_func_name)
        billing_display.pack(side='left')

def supplierFunction():
    global func_list
    prev_func_name = func_list[0]
    current_func_name = "supplier"
    if prev_func_name == current_func_name:
        pass
    else:
        func_finder(prev_func_name)
        func_list.insert(0, current_func_name)
        supply_display.pack(side='left')

def historyFunction():
    global func_list
    prev_func_name = func_list[0]
    current_func_name = "history"
    if prev_func_name == current_func_name:
        pass
    else:
        func_finder(prev_func_name)
        func_list.insert(0, current_func_name)
        history_display.pack(side='left')
        show_local_history_data()
        show_global_history_data()

def settingFunction():
    global func_list
    prev_func_name = func_list[0]
    current_func_name = "setting"
    if prev_func_name == current_func_name:
        pass
    else:
        func_finder(prev_func_name)
        func_list.insert(0, current_func_name)
        setting_display.pack(side='left')

def ccFunction():
    global func_list
    prev_func_name = func_list[0]
    current_func_name = "cc"
    if prev_func_name == current_func_name:
        pass
    else:
        func_finder(prev_func_name)
        func_list.insert(0, current_func_name)
        cc_display.pack(side='left')

def stock_add():
    global inventory_func_list
    inventory_display.pack_forget()
    add_stock_frame.pack(side='left')
    
    inventory_func_list.insert(0, 'add_stock')

def update_stock():
    global inventory_func_list
    inventory_display.pack_forget()   
    update_stock_frame.pack(side='left')
    inventory_func_list.insert(0, 'update_stock')

def analyse_stock():
    global inventory_func_list
    inventory_display.pack_forget()   
    stock_analytics_frame.pack(side='left')
    inventory_func_list.insert(0, 'analyse_stock')
    fetch_and_display_inventory()

def open_calendar():
    top = Toplevel(win)
    top.title("Select a Date")
    top.geometry("300x300")

    now_spec = datetime.now()

    # Extract year, month, and day
    YEAR = now_spec.year
    MONTH = now_spec.month
    DATE = now_spec.day
    
    cal = Calendar(top, selectmode='day', year=YEAR, month=MONTH, day= DATE)
    cal.pack(pady=10)

    def grab_date():
        selected_date = cal.get_date()  # Format is typically mm/dd/yyyy
        date_obj = datetime.strptime(selected_date, "%m/%d/%y")  # Parse input
        formatted_date = date_obj.strftime("%d/%m/%Y")   
        mfdate.set(formatted_date)
        # date_label.configure(text=f"Date: {selected_date}")
        top.destroy()

    ctk.CTkButton(top, text="Select", command=grab_date).pack(pady=5)
    
def open_calendar1():
    top = Toplevel(win)
    top.title("Select a Date")
    top.geometry("300x300")

    now_spec = datetime.now()

    # Extract year, month, and day
    YEAR = now_spec.year
    MONTH = now_spec.month
    DATE = now_spec.day
    
    cal = Calendar(top, selectmode='day', year=YEAR, month=MONTH, day= DATE)
    cal.pack(pady=10)

    def grab_date():
        selected_date = cal.get_date()  # Format is typically mm/dd/yyyy
        date_obj = datetime.strptime(selected_date, "%m/%d/%y")  # Parse input
        formatted_date = date_obj.strftime("%d/%m/%Y")
        expdate.set(formatted_date)
        # date_label.configure(text=f"Date: {selected_date}")
        top.destroy()

    ctk.CTkButton(top, text="Select", command=grab_date).pack(pady=5)

def add_stock_to_dbs():
    global inventory
      
    find_product_id = inventory.find_one(
    {"Product_ID": barid.get()})
    
    find_product_name = inventory.find_one(
        {"Product_Name": productname.get().lower()}
    )
    
    if find_product_id or find_product_name:
        return "0"
    else:
        return "1"
        

def add_stock():
    global inventory
    if len(barid.get())>0:
        eligibility = add_stock_to_dbs()

        if eligibility is True:
            print("True")
        elif eligibility is False:
            print("False")
        else:
            if eligibility == '0':
                response = messagebox.askyesno("Confirm", "The Product with the given ID ot Name already exist. Do you want to update stock Quantity?")
                if response:
                    check_quantity = inventory.find_one({'Product_ID':barid.get()}, {'Quantity':1})
                    if check_quantity:
                        if check_quantity['Quantity'] == 0:
                            try:
                                if len(mfdate.get())>=4 and len(expdate.get())>= 4 and int(productqty.get()) > 0:
                                    product_dict_to_update = {
                                        'Quantity': int(productqty.get()),
                                        'Manufacture_Date': mfdate.get(),
                                        'Expire_Date': expdate.get()
                                    }
                                    
                                    inventory.update_one({'Product_ID':barid.get()}, {"$set":product_dict_to_update})
                                    
                                    find_product_name_lc = inventory.find_one(
                                        {"Product_ID": barid.get()}, {"Product_Name":1})
                                    find_product_cp_lc = inventory.find_one(
                                        {"Product_ID": barid.get()}, {"Cost_Price":1})
                                    
                                    localhistory_update(product_id1 = barid.get(), product_name1 = find_product_name_lc, amount1 = find_product_cp_lc*float(productqty.get()), action1 = int(productqty.get()))
                                    globalhistory_update(product_id1s = barid.get(), product_name1s = find_product_name_lc ,amount1s = find_product_cp_lc*float(productqty.get()) , action1s = int(productqty.get()))

                                    messagebox.showinfo('Stock Update', f"Stock with ID {barid.get()} is updated!")
                                else:
                                    messagebox.showinfo('Stock Quantity Update', 'While updating the stock quantity, you need to re-enter Manufacture and Expire date and product quantity.')
                            except ValueError:
                                messagebox.showerror('')
                                
                        else:
                            messagebox.showinfo('Product Quantity', f"You can not add more stock as you have {str(check_quantity['Quantity'])} still left!Add when it is Zero!")
                    else:
                        messagebox.showerror('ID Error', 'Unexpected error occured!')
                else:
                    messagebox.showwarning("Duplication", "Product name or Product ID already exist!")
                    should_add = False
            else:
                should_add = True
                if len(productname.get())>0  and len(productqty.get())>0 and len(cp.get())>0 and len(sp.get())>0 and len(mfdate.get())>=4 and len(expdate.get())>= 4:
                    try:
                        if int(productqty.get()) < 0:
                            messagebox.showerror('Product Quantity','Invalid Product Quantity')
                            should_add = False
                        elif float(cp.get()) < 0:
                            messagebox.showerror('Cost Price','Invalid Cost Price!')
                            should_add = False
                        elif float(sp.get()) < 0:
                            messagebox.showerror('Selling Price','Invalid Selling Price!')
                            should_add = False
                    except ValueError:
                        messagebox.showerror('Input Error', 'Product Quantity must be an integer!\nor\nSP or CP must be Positive Number')
                        should_add = False
                else:
                    messagebox.showinfo('Empty Inputs', 'Entry fields must not be empty!')
                    should_add = False
                    
                if should_add is not False:      
                    should_add = True
                    if len(tax.get()) == 0:
                        taxs = float(0)
                    else:
                        try:    
                            taxs = float(tax.get())
                        except ValueError:
                            messagebox.showerror('Input Error','Tax Entry Should not contain alphabets or special Characters!')
                            should_add = False
        try:                
            if should_add is True:
                print("Entering This area")
                product_dict_to_add = {
                    'Product_ID': barid.get(),
                    'Product_Name': productname.get().lower(),
                    'Quantity': int(productqty.get()),
                    'Cost_Price': float(cp.get()),
                    'Selling_Price': float(sp.get()),
                    'Tax': taxs,
                    'Discount': 0.0,
                    'Manufacture_Date': mfdate.get(),
                    'Expire_Date': expdate.get()
                }
                # print(product_dict_to_add)
                inventory.insert_one(product_dict_to_add)
                
                localhistory_update(product_id1 = barid.get(), product_name1 = productname.get().lower(), amount1 = float(cp.get())*float(productqty.get()), action1 = int(productqty.get()))
                globalhistory_update(product_id1s = barid.get(), product_name1s = productname.get().lower() ,amount1s = float(cp.get())*float(productqty.get()) , action1s = int(productqty.get()))

                messagebox.showinfo('Successfull', 'Product Added Successfully!')
            else:
                pass
        except UnboundLocalError:
            pass
    else:
        messagebox.showinfo('ID not found', 'Please Scan the product barcode!')
 
def clear_inventory_entries():
    bar_id_entry.delete(0, 'end')
    productname_entry.delete(0, 'end')
    productQuantity_entry.delete(0, 'end')
    cp_entry.delete(0, 'end')
    sp_entry.delete(0, 'end')
    tax_entry.delete(0, 'end')
    mf_date_entry.delete(0, 'end')
    exp_date_entry.delete(0, 'end')

def back_stock_entry():
    add_stock_frame.pack_forget()
    inventory_display.pack(side='left')
def back_update_entry():
    update_stock_frame.pack_forget()
    inventory_display.pack(side='left')
def analytics_back():
    stock_analytics_frame.pack_forget()
    inventory_display.pack(side='left')
    
def find_item_in_update():
    global product_id_found
    id_val = item_id_in_update.get()
    find_product = inventory.find_one(
        {"Product_ID": id_val}
    )
    try:
        if find_product:
            find_result_label.configure( text = f"Item Found:{find_product['Product_Name']}", fg ='White', bg ='#4B54F8', width=(len(find_product)+10) )
            find_result_label.place(x = 585, y = 180) #(x = 585, y = 118)
            product_id_found['Update'] = True
            product_id_found['UID'] = find_product['Product_ID']
        else:
            find_result_label.configure( text = "Item Not Found!",fg ='White', bg ='red', width=15)
            find_result_label.place(x = 585, y = 180)
            product_id_found['Update'] = False
            product_id_found['UID'] = find_product['Product_ID']
    except TypeError:
        pass

def update_product():
    global product_id_found
    should_update = product_id_found['Update']
    update_ID = product_id_found['UID']
    if should_update is True:
        try:
            if len(newproductval.get()) != 0:
                inventory.update_one({'Product_ID': update_ID}, {"$set":{'Product_Name':newproductval.get().strip()}})
            if len(newcpval.get()) != 0:
                inventory.update_one({'Product_ID': update_ID}, {"$set":{'Cost_Price':float(newcpval.get())}})
            if len(newspval.get()) != 0:
                inventory.update_one({'Product_ID': update_ID}, {"$set":{'Selling_Price':float(newspval.get())}})
            if len(newtaxval.get()) != 0:
                inventory.update_one({'Product_ID': update_ID}, {"$set":{'Tax':float(newtaxval.get())}})
            if len(newdiscountval.get()) != 0:
                inventory.update_one({'Product_ID': update_ID}, {"$set":{'Discount':float(newdiscountval.get())}})
            else:
                messagebox.showinfo('Update', 'Updated!')
        except ValueError:
            messagebox.showerror('Invalid Input/s', 'Cost Price, Selling Price, Tax and Discount should not be any alphabet and should be greater than zero')
    else:
        messagebox.showerror('ID Error', 'ID Not Found!')
    
    
def delete_update_entries():
    bar_entry.delete(0, 'end')
    new_product_entry.delete(0, 'end')
    new_cp_entry.delete(0, 'end')
    new_sp_entry.delete(0, 'end')
    new_tax_entry.delete(0, 'end')
    new_discount_entry.delete(0, 'end')
    find_result_label.place_forget()

def fetch_and_display_inventory():
    analytics_board.config(state=NORMAL)
    analytics_board.delete("1.0", END)

    # Fetch data from MongoDB and create DataFrame
    data = list(inventory.find())
    if not data:
        analytics_board.insert(END, "No data found in the database.")
        return

    df = pd.DataFrame(data)

    # Define columns and ensure they exist
    columns = [
        "SLNO", "Product_ID", "Product_Name", "Cost_Price", "Selling_Price",
        "Manufacture_Date", "Expire_Date", "Tax", "Discount"
    ]

    for col in columns[1:]:  # excluding SLNO
        if col not in df.columns:
            df[col] = ""

    df.insert(0, "SLNO", range(1, len(df) + 1))
    df = df[columns].fillna("")

    # Column widths
    col_widths = {
        "SLNO": 5,
        "Product_ID": 12,
        "Product_Name": 20,
        "Cost_Price": 12,
        "Selling_Price": 14,
        "Manufacture_Date": 18,
        "Expire_Date": 15,
        "Tax": 6,
        "Discount": 9
    }

    # Format lines
    separator = "|"
    format_str = separator.join(f"{{:<{col_widths[col]}}}" for col in columns)
    divider_line = "-" * (sum(col_widths.values()) + len(columns) - 1)

    # Header
    analytics_board.insert(END, format_str.format(*columns) + "\n")
    analytics_board.insert(END, divider_line + "\n")

    # Rows with dividers
    for _, row in df.iterrows():
        row_line = format_str.format(*[str(row[col]) for col in columns])
        analytics_board.insert(END, row_line + "\n")
        analytics_board.insert(END, divider_line + "\n")

    analytics_board.config(state=DISABLED)
    
def analytics_idsrc_display():
    analytics_board.config(state=NORMAL)
    analytics_board.delete("1.0", END)
    find_product_id = inventory.find_one(
        {'Product_ID':analytics_src_val.get()}
    )
    if find_product_id:
        product_dict = inventory.find_one(
            {'Product_ID':analytics_src_val.get()},
            {"Product_Name": 1,
             "Cost_Price": 1,
             "Selling_Price": 1,
             "Manufacture_Date": 1,
             "Expire_Date": 1,
             "Tax": 1,
             "Discount": 1}
        )
        arranged_product = f'''
#######---- Product Found ----#######

Product_ID:       {analytics_src_val.get()},
Product_Name:     {product_dict['Product_Name']},
Cost_Price:       {product_dict['Cost_Price']},
Selling_Price:    {product_dict['Selling_Price']},
Manufacture_Date: {product_dict['Manufacture_Date']},
Expire_Date:      {product_dict['Expire_Date']},
Tax:              {product_dict['Tax']},
Discount:         {product_dict['Discount']}
        '''
        analytics_board.insert(END, arranged_product)
    else:
        messagebox.showerror('Wrong ID', 'Product not found!')

# analytics_option = ['All', 'Cost, Selling Price', 'Product, Selling Price','Product, Cost Price', 'Product, Discount', 'Product, Tax', 'Product, Tax, Discount']

def fetch_and_display_analytics(selected_columns):
    analytics_board.config(state=NORMAL)
    analytics_board.delete("1.0", END)

    # Fetch data from MongoDB and create DataFrame
    data = list(inventory.find())
    if not data:
        analytics_board.insert(END, "No data found in the database.")
        return

    df = pd.DataFrame(data)

    # Ensure selected columns exist in DataFrame
    valid_columns = []
    for col in selected_columns:
        if col not in df.columns:
            df[col] = ""  # Add missing columns as empty if needed
        valid_columns.append(col)

    # Add SLNO at the beginning
    df.insert(0, "SLNO", range(1, len(df) + 1))
    if "SLNO" not in valid_columns:
        valid_columns.insert(0, "SLNO")

    df = df[valid_columns].fillna("")

    # Define default column widths (can be customized)
    length_list = []
    for length in selected_columns:
        length_list.append(len(length))

    default_width = max(length_list)+5
    col_widths = {
        col: (5 if col == "SLNO" else max(default_width, len(col)))
        for col in valid_columns
    }

    # Format lines
    separator = "|"
    format_str = separator.join(f"{{:<{col_widths[col]}}}" for col in valid_columns)
    divider_line = "-" * (sum(col_widths.values()) + len(valid_columns) - 1)

    # Header
    analytics_board.insert(END, format_str.format(*valid_columns) + "\n")
    analytics_board.insert(END, divider_line + "\n")

    # Rows
    for _, row in df.iterrows():
        row_line = format_str.format(*[str(row[col]) for col in valid_columns])
        analytics_board.insert(END, row_line + "\n")
        analytics_board.insert(END, divider_line + "\n")

    analytics_board.config(state=DISABLED)


def option_selected(choice):
    global choice_list
    if choice == 'All':
        fetch_and_display_inventory()
        choice_list.insert(0, 'All')

    elif choice == 'Cost, Selling Price':
        print('Choice 2')
        fetch_and_display_analytics(['Product_ID','Product_Name', 'Cost_Price', 'Selling_Price'])
        choice_list.insert(0, 'Cost, Selling Price')
        
    elif choice == 'Product, Selling Price':
        fetch_and_display_analytics(['Product_ID','Product_Name', 'Selling_Price'])
        choice_list.insert(0, 'Product, Selling Price')

    elif choice == 'Product, Cost Price':
        fetch_and_display_analytics(['Product_ID','Product_Name', 'Cost_Price'])
        choice_list.insert(0, 'Product, Cost Price')

    elif choice == 'Product, Discount':
        fetch_and_display_analytics(['Product_ID','Product_Name', 'Discount'])
        choice_list.insert(0, 'Product, Discount')
    elif choice == 'Product, Tax':
        fetch_and_display_analytics(['Product_ID','Product_Name', 'Tax'])
        choice_list.insert(0, 'Product, Tax')
    elif choice == 'Product, Tax, Discount':
        fetch_and_display_analytics(['Product_ID','Product_Name', 'Tax', 'Discount'])
        choice_list.insert(0, 'Product, Tax, Discount')
    else:
        messagebox.showerror('Invalid Option', 'Option not found!')


def export_inventory_to_excel(selected_columns):

    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        title="Save Inventory Data As"
    )

    if not file_path:
        print("Export cancelled by user.")
        return

    # Fetch data from MongoDB
    data = list(inventory.find())
    if not data:
        print("No data found in the database.")
        return

    df = pd.DataFrame(data)

    # Ensure selected columns exist
    valid_columns = []
    for col in selected_columns:
        if col not in df.columns:
            df[col] = ""
        valid_columns.append(col)

    # Add SLNO
    df.insert(0, "SLNO", range(1, len(df) + 1))
    if "SLNO" not in valid_columns:
        valid_columns.insert(0, "SLNO")

    df = df[valid_columns].fillna("")

    # Save to Excel
    try:
        df.to_excel(file_path, index=False)

        # Load workbook and format
        wb = load_workbook(file_path)
        ws = wb.active

        center_alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, col_name in enumerate(valid_columns, start=1):
            col_letter = get_column_letter(col_idx)
            if col_name == "SLNO":
                ws.column_dimensions[col_letter].width = 6
            elif len(col_name) > 15:
                ws.column_dimensions[col_letter].width = len(col_name) + 5
            else:
                ws.column_dimensions[col_letter].width = 15

            for cell in ws[col_letter]:
                cell.alignment = center_alignment

        wb.save(file_path)
        messagebox.showinfo('Excel File Saved',f"Excel file saved: {file_path}")

        # Convert Excel to PDF
        pdf_path = os.path.splitext(file_path)[0] + ".pdf"
        pythoncom.CoInitialize()  # Initialize COM (required in some contexts)
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb_excel = excel.Workbooks.Open(file_path)
        wb_excel.ExportAsFixedFormat(0, pdf_path)
        wb_excel.Close(False)
        excel.Quit()
        messagebox.showinfo('PDF Saved',f"PDF file saved: {pdf_path}")

    except Exception as e:
        messagebox.showerror('Conversion Failed!',f"Failed to save or convert file: {e}")


def save_analytics_display():
    global choice_list
    try:
        if choice_list[0] == 'All':
            
            col = [
                "Product_ID", "Product_Name", "Cost_Price", "Selling_Price",
                "Manufacture_Date", "Expire_Date", "Tax", "Discount"
            ]
            export_inventory_to_excel(col)
            
        elif choice_list[0] == 'Cost, Selling Price':
            export_inventory_to_excel(['Product_ID','Product_Name', 'Cost_Price', 'Selling_Price'])
            
        elif choice_list[0] == 'Product, Selling Price':
            export_inventory_to_excel(['Product_ID','Product_Name', 'Selling_Price'])

        elif choice_list[0] == 'Product, Cost Price':
            export_inventory_to_excel(['Product_ID','Product_Name', 'Cost_Price'])

        elif choice_list[0] == 'Product, Discount':
            export_inventory_to_excel(['Product_ID','Product_Name', 'Discount'])
        elif choice_list[0] == 'Product, Tax':
            export_inventory_to_excel(['Product_ID','Product_Name', 'Tax'])
        elif choice_list[0] == 'Product, Tax, Discount':
            export_inventory_to_excel(['Product_ID','Product_Name', 'Tax', 'Discount'])
        else:
            messagebox.showerror('Invalid Option', 'Option not found!')
    except IndexError:
        messagebox.showerror('IndexError', 'No parameter Found!')

def get_cc_text():
    global username_list
    print(username_list[0])
    content_cc = cc_textbox.get("0.0", "end")  # "0.0" = start, "end-1c" = end minus last newline
    if len(content)>0:
        now = datetime.now()
        date_time_cc = now.strftime("%Y-%m-%d %H:%M:%S")
        email = client_info.find_one({'username':username_list[0]}, {'email':1})
        if email:
            Email = email['email']
            cc_message = {
                'Email':Email,
                'Time': date_time_cc,
                'Message': content_cc
            }
            cc_database.insert_one(cc_message)
            messagebox.showinfo('Customer Care', f'Message sent successfully.The reply will sent to your email {Email}')
        else:
            messagebox.showerror('Customer Care', "Can't send the message! Please login again.")
    else:
        messagebox.showerror('Customer Care', "Can't send empty message!")
def insert_in_alert_textbox(idee):
    show_data = inventory.find_one(
        {'Product_ID':idee},
        {"Product_Name": 1,
         "Quantity":1,
        "Cost_Price": 1,
        "Selling_Price": 1,
        "Manufacture_Date": 1,
        "Expire_Date": 1,
        "Tax": 1,
        "Discount": 1}
    )
    data_insert_alert_txtbx = f'''
Product_ID:       {idee},
Product_Name:     {show_data['Product_Name']},
Quantity:         {show_data['Quantity']},  
Cost_Price:       {show_data['Cost_Price']},
Selling_Price:    {show_data['Selling_Price']},
Manufacture_Date: {show_data['Manufacture_Date']},
Expire_Date:      {show_data['Expire_Date']},
Tax:              {show_data['Tax']},
Discount:         {show_data['Discount']}
            
            '''
    alert_textbox.insert("0.0", data_insert_alert_txtbx)
            
def on_row_selected_expire(value):
    alert_textbox.delete("0.0", 'end')
    selected = alert_tree_expire.focus()
    if selected:
        values = alert_tree_expire.item(selected, "values")
        if values:
            product_id = values[1]
            insert_in_alert_textbox(product_id)

def on_row_selected_stock(value):
    alert_textbox.delete("0.0", 'end')
    selected = alert_tree_stock.focus()
    if selected:
        values = alert_tree_stock.item(selected, "values")
        if values:
            product_id = values[1]
            insert_in_alert_textbox(product_id)

def insert_data_to_alert_treeview():
    
    for items in alert_tree_expire.get_children():
        alert_tree_expire.delete(items)
        
    alert_tree_expire.tag_configure('expiring_soon', foreground='red')
    alert_tree_expire.tag_configure('expiring_late', foreground='green')
    inventory_data = inventory.find({})
    count = 1
    
    date_format = "%d/%m/%Y"
    date1 = datetime.strptime(datetime.today().strftime(date_format), date_format)
    
    for data in inventory_data:
        try:
            expire_date_db = data['Expire_Date']
            date2 = datetime.strptime(expire_date_db, date_format)
            
            date_diff = abs((date2 - date1).days)

            if date_diff<29:
                item = (count, data['Product_ID'], data['Product_Name'], str(date_diff)+" days")
                alert_tree_expire.insert("", "end", values=item, tags=('expiring_soon',) )
                count = count+1
            else:
                # item = (count, data['Product_ID'], data['Product_Name'], str(date_diff)+" days")
                # alert_tree_expire.insert("", "end", values=item, tags=('expiring_late',))
                # count = count+1
                pass
        except (ValueError, ZeroDivisionError):
            pass
        
        
def insert_data_to_alert_treeview_stock():
    inventory_data = inventory.find({})
    for items in alert_tree_stock.get_children():
        alert_tree_stock.delete(items)
    
    alert_tree_stock.tag_configure('stock_low', foreground='red')
    alert_tree_stock.tag_configure('stock_ok', foreground='green')
    
    count2 = 1
    for  stockk in inventory_data:
        if stockk['Quantity'] <= 20:
            item = (count2, stockk['Product_ID'], stockk['Product_Name'], stockk['Quantity'])
            alert_tree_stock.insert("", "end", values=item, tags=('stock_low',) )
            count2 = count2+1
        else:
            # item = (count2, stockk['Product_ID'], stockk['Product_Name'], stockk['Quantity'])
            # alert_tree_stock.insert("", "end", values=item, tags=('stock_ok',))
            # count2 = count2+1
            pass

            
def clear_alert_texbox():
    alert_textbox.delete("0.0", "end")
    
click_count = 1
def switch_to_stock():
    global click_count
    if click_count%2 != 0:
        # print("odd")
        exchange_label_button.configure(text = 'Show Expire Alert')
        click_count = click_count + 1
        expiring_frame.pack_forget()
        stock_alert_frame.pack(side='left', anchor='nw', padx=20, pady=(170, 0))
        
        insert_data_to_alert_treeview_stock()
    else:
        # print("even")
        exchange_label_button.configure(text = 'Show Stock Alert')
        click_count = click_count + 1
        stock_alert_frame.pack_forget()
        expiring_frame.pack(side='left', anchor='nw', padx=20, pady=(170,0))

def CGST_rs(pdct_price, pdct_qty, cgst):
    if cgst == 0 :
        return 0
    else:
        return (pdct_qty*pdct_price*cgst)/100
    
def SGST_rs(pdct_price, pdct_qty, sgst):
    if sgst == 0:
        return 0
    else:
        return((pdct_qty*pdct_price*sgst)/100)
        
def total_amount(pdct_price, pdct_qty, sgst, cgst, discount):
    if discount == 0:
        return ((pdct_price*pdct_qty)+sgst+cgst)
    else:
        discount_value = (((pdct_price*pdct_qty)+sgst+cgst))*discount/100
        actual_amount = ((pdct_price*pdct_qty)+sgst+cgst)
        return actual_amount - discount_value
def check_product_quantity(pdct_qty, pdct_id):
    find_quantity = inventory.find_one({"Product_ID":pdct_id},{"Quantity":1})
    if find_quantity["Quantity"] >= pdct_qty+1:
        return True
    else:
        return False
           
billing_slno = 1
product_ids = []
def billing_tree_insert(value):
    global billing_slno, product_ids

    inv_data = inventory.find_one({'Product_ID':barcodevalue.get()})
    
    if inv_data:
        product_dict_bill = inventory.find_one(
            {'Product_ID':barcodevalue.get()},
            {"Product_Name": 1,
            "Selling_Price": 1,
            "Tax": 1,
            "Discount": 1}
        )

        for item0 in billing_tree.get_children():
            values0 = billing_tree.item(item0)["values"]
            product_ids.append(str(values0[1]))  # Product ID is at index 1
            
        if barcodevalue.get() in product_ids:
            
            for item in billing_tree.get_children():
                values = list(billing_tree.item(item)["values"])
                if str(values[1]) == barcodevalue.get():  # Product ID is at index 1
                    product_current_qty = values[3]
                    chk_quantity = check_product_quantity(product_current_qty, barcodevalue.get())
                    if chk_quantity is True:
                        cgst_rs = CGST_rs(product_dict_bill["Selling_Price"], product_current_qty+1, product_dict_bill["Tax"]/2)
                        sgst_rs = CGST_rs(product_dict_bill["Selling_Price"], product_current_qty+1, product_dict_bill["Tax"]/2)
                        amount_total = total_amount(product_dict_bill["Selling_Price"], product_current_qty+1,sgst_rs, cgst_rs, product_dict_bill["Discount"])
                        values[3] = int(product_current_qty+1)  # Quantity is at index 3
                        values[6] = sgst_rs
                        values[8] = cgst_rs
                        values[10] = amount_total
                        billing_tree.item(item, values=values)
                    else:
                        messagebox.showinfo('Stock Alert', f"Stock Exhausted for {values[2]}")
        else:
            init_chk_quantity = check_product_quantity(0, barcodevalue.get())
            if init_chk_quantity is True:
                cgst_rs = CGST_rs(product_dict_bill["Selling_Price"], 1, product_dict_bill["Tax"]/2)
                sgst_rs = CGST_rs(product_dict_bill["Selling_Price"], 1, product_dict_bill["Tax"]/2)
                amount_total = total_amount(product_dict_bill["Selling_Price"], 1,sgst_rs, cgst_rs, product_dict_bill["Discount"])
                
                product_data = [
                    (billing_slno, barcodevalue.get(), product_dict_bill["Product_Name"], 1, product_dict_bill["Selling_Price"],
                    product_dict_bill["Tax"]/2, sgst_rs, product_dict_bill["Tax"]/2, cgst_rs, product_dict_bill["Discount"], amount_total)
                ]
                billing_tree.insert("", END, values=product_data[0])
                billing_slno = billing_slno+1
            else:
                messagebox.showinfo("Stock Alert", f"Stock Exhausted for {product_dict_bill['Product_Name']}")
            
        barcode_id_entry.delete(0, "end")
    else:
        messagebox.showerror("ID Error", "Product ID not found")
  
def clear_bill_treeview():
    global billing_slno, product_ids
    for item in billing_tree.get_children():
        billing_tree.delete(item)
        
    billing_slno = 1
    product_ids = []  
    
payment_method = "CASH"

def on_cash_click():
    global payment_method
    if upi_var.get():
        upi_var.set(0)
        payment_method = "CASH"
def on_upi_click():
    global payment_method
    if cash_var.get():
        cash_var.set(0)
        payment_method = "UPI"
def on_checkbox2_click():
    if def_var.get():
        def_var.set(0)

def on_checkbox1_click():
    if a4_var.get():
        a4_var.set(0)



def gen_bill_win():
    data_dict = {}
    global payment_method
    
    def save_and_print():
        global payment_method
        filename = f"Bill_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        full_path = os.path.join(save_path, filename)

        paper_width = 80 * mm
        paper_height = 200 * mm
        c = canvas.Canvas(full_path, pagesize=(paper_width, paper_height))

        x_margin = 5 * mm
        y_position = paper_height - 10 * mm
        line_height = 10

        c.setFont("Courier", 8)

        for line in bill_text.split('\n'):
            c.drawString(x_margin, y_position, line)
            y_position -= line_height
            if y_position < 10:
                c.showPage()
                c.setFont("Courier", 8)
                y_position = paper_height - 10 * mm

        c.save()
        messagebox.showinfo("Bill Saved", "Bill Saved Successfully.")
        
        for pdid, itemm in data_dict.items():
            result_found = inventory.update_one(
                {"Product_ID": str(pdid)},
                {"$inc": {"Quantity": -int(itemm["Quantity"])}}
            )
            if result_found.matched_count <= 0:
                print("Can't update the db")
        
            localhistory_update(product_id1 = pdid, product_name1 = itemm["Product Name"],amount1 = itemm["Amount"], action1 = -int(itemm["Quantity"]))
            globalhistory_update(product_id1s = pdid, product_name1s = itemm["Product Name"], amount1s = itemm["Amount"], action1s = -int(itemm["Quantity"]))

        update_earnings()
        webbrowser.open(f"file://{os.path.abspath(full_path)}")
        bill_win.destroy()
            
    bill_win = Toplevel(win)
    bill_win.title("Generate Invoice")
    
    save_path = r"./bills"
    os.makedirs(save_path, exist_ok=True)
    
    if def_checkbox.get() == 1:
        bill_win.geometry("430x700+310+30")
        bill_text_box = Text(bill_win, height=10, width=40, wrap="none", relief='ridge', bd=2)
        bill_text_box.pack(side="left", fill="both", expand=True, pady=(0, 60))

        # Create the Scrollbar and link it to the Text widget
        scrollbar = Scrollbar(bill_win, command=bill_text_box.yview)
        scrollbar.pack(side="right", fill="y")

        # Configure the Text widget to update with the scrollbar
        bill_text_box.config(yscrollcommand=scrollbar.set)
        for data in billing_tree.get_children():
            datas = billing_tree.item(data)["values"]
            data_dict[datas[1]] = {"SLNO":datas[0], "Product Name":datas[2],
                                   "Quantity":datas[3], "MRP":datas[4],"Amount": datas[10]}

        print(data_dict)
        try:
            with open(file_path, 'r') as file:
                content = file.read().strip()
                
        except FileNotFoundError:
            messagebox.showerror('File Error',f"The file '{file_path}' was not found.")
        except IOError:
            messagebox.showerror('Read Error',f"An error occurred while reading the file '{file_path}'.")

        df_format = pd.DataFrame.from_dict(data_dict, orient='index')

        # Compute total
        try:
            total_amount = df_format['Amount'].astype(float).sum()
        except KeyError:
            total_amount = 0

        # Create formatted string
        pre_header = f"Shop : {content} \nAddress : A.S College Road,Bikramganj\nMob : 9608053244\nEmail : yourajverma960@gmail.com\nGSTIN : 10CDLP1ZQ\n==================== Buyer Details ====================\n"
        header = "==================== INVOICE ====================\n"
        footer = f"\n------------------------------------------------\nTotal Amount: {total_amount:.2f}\n================================================\nThank you for your purchase!\n"
        payment = f"Payment Method : {payment_method}\n"
        today = datetime.today().strftime('%d/%m/%Y %H:%M:%S')
        
        buyer_detail = ""
        if buyer_name_entry.get():
            buyer_detail += f"Buyer Name : {buyer_name_entry.get()}\n"
        if buyer_phone_entry.get():
            buyer_detail += f"Buyer Phone : {buyer_phone_entry.get()}\n"
        if buyer_address_entry.get():
            buyer_detail += f"Buyer Add : {buyer_address_entry.get()}\n"
        else:
            buyer_detail = "NONE\n"
        # Format DataFrame for text display
        table_str = df_format.to_string(index=False)

        # Final bill text
        bill_text =pre_header + buyer_detail + header + table_str + footer + payment + today
        
        bill_text_box.insert(END, bill_text)
        
        save_print_btn = ctk.CTkButton(bill_win, font=('Poppins', 18, 'bold'), text="Generate Bill", height=35, width = 200, fg_color="#28F13F",
                                     bg_color="white", corner_radius=13, hover_color='black', text_color="white",
                                     command = save_and_print)
        save_print_btn.place(x = 100, y = 650)
                
    else:
        pass
    
current_dir_his = Path(__file__).resolve().parent if "__file__" in locals() else Path.cwd()

def localhistory_update(product_id1, product_name1,amount1, action1):
    global current_dir_his
    
    now = datetime.now()


    date_str = now.strftime("%d/%m/%Y")


    time_str = now.strftime("%H:%M:%S")

    db_path = current_dir_his/"Data"/"localhistory.db"
    conn_lc = sqlite3.connect(db_path)

    cursor_lc = conn_lc.cursor()

    create_table_query = '''
    CREATE TABLE IF NOT EXISTS lchistory (
        Product_ID TEXT,
        Product_Name TEXT,
        Date TEXT,
        Time TEXT,
        Amount REAL,
        Action INTEGER
    )
    '''
    cursor_lc.execute(create_table_query)
    
    product_id = product_id1
    product_name = product_name1
    date = date_str
    time = time_str
    amount = amount1
    action = action1

    # SQL insert statement
    insert_query = '''
    INSERT INTO lchistory (Product_ID, Product_Name, Date, Time, Amount, Action)
    VALUES (?, ?, ?, ?, ?, ?)
    '''

    cursor_lc.execute(insert_query, (product_id, product_name, date, time, amount, action))

    conn_lc.commit()
    conn_lc.close()


    
def globalhistory_update(product_id1s, product_name1s,amount1s, action1s):
    global username_list, current_dir_his
    
    try:
        email = client_info.find_one({'username':username_list[0]}, {'email':1})
        
        if not email:
            email = "None"
    except IndexError:
        email = "None"
    
    now = datetime.now()


    date_str = now.strftime("%d/%m/%Y")


    time_str = now.strftime("%H:%M:%S")
    
    glob_history_dict = {
        "product_id" : product_id1s,
        "Email" : email,
        "product_name" : product_name1s,
        "date" : date_str,
        "time" : time_str,
        "amount" : amount1s,
        "action" : action1s
    }

    global_history.insert_one(glob_history_dict)

        
def show_local_history_data():
    global username_list, current_dir_his
    
    for item in history_tree.get_children():
        history_tree.delete(item)
    
    db_path = current_dir_his/"Data"/"localhistory.db"
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row  # Enables dict-like access

    cursor = conn.cursor()

    # Select all rows from the table
    cursor.execute('SELECT * FROM lchistory')
    rows = cursor.fetchall()

    # Convert each row to a dictionary
    data = []
    for row in rows:
        row_dict = dict(row)
        data.append(row_dict)

    conn.close()
    
    history_tree.tag_configure('sell', foreground="#f92d2d")  
    history_tree.tag_configure('buy', foreground="#229954")

    # Example output
    for row in data:
        history_data_format = (row["Product_ID"], row["Product_Name"], row["Date"], row["Time"], row["Amount"], row["Action"])
        if int(row["Action"]) < 0:
            history_tree.insert("", END, values = history_data_format, tags=("sell",))
        else:
            history_tree.insert("", END, values = history_data_format, tags=("buy",))

def show_global_history_data():
    for item in glo_history_tree.get_children():
        glo_history_tree.delete(item)
        
    glob_history_data = global_history.find()

    glo_history_tree.tag_configure('sell', foreground="#f92d2d")  
    glo_history_tree.tag_configure('buy', foreground="#229954")
    
    # Print each document
    for doc in glob_history_data:
        history_data_format2 = (doc["product_id"], doc["Email"], doc["product_name"], doc["date"], doc["time"], doc["amount"], doc["action"])
        if int(doc["action"]) < 0:
            glo_history_tree.insert("", END, values = history_data_format2, tags=("sell",))
        else:
            glo_history_tree.insert("", END, values = history_data_format2, tags=("buy",))


def show_local_history():
    global_history_frame.pack_forget()
    local_history_frame.pack(side='bottom', anchor='s', fill=X, expand=True, padx=10, pady=(0, 10))
    show_local_history_data()

def show_global_history():
    local_history_frame.pack_forget()
    global_history_frame.pack(side='bottom', anchor='s', fill=X, expand=True, padx=10, pady=(0, 10))
    show_global_history_data()

def get_week_dates(date_str):
    # Convert the string to a datetime object
    given_date = datetime.strptime(date_str, "%d/%m/%Y")

    # Find the Monday of the current week
    start_of_week = given_date - timedelta(days=given_date.weekday())

    # Generate all dates from Monday to Sunday
    week_dates = [(start_of_week + timedelta(days=i)).strftime("%d/%m/%Y") for i in range(7)]
    return week_dates
    

def get_current_week_of_month():
    import datetime
    today = datetime.date.today()
    year = today.year
    month = today.month

    # First day of the month
    first_day = datetime.date(year, month, 1)

    # Find first Monday in the month
    first_day_weekday = first_day.weekday()  # 0 = Monday
    days_to_monday = (7 - first_day_weekday) % 7
    first_monday = first_day if first_day_weekday == 0 else first_day + datetime.timedelta(days=days_to_monday)

    # Calculate week number
    if today < first_monday:
        week_number = 1
    else:
        delta_days = (today - first_monday).days
        week_number = (delta_days // 7) + 2

    # Get short month name
    month_short = today.strftime("%b")

    return week_number, month_short

sum_amount_day = 0
sum_amount_week = 0
sum_amount_month = 0

def update_earnings():
    global sum_amount_day, sum_amount_month, sum_amount_week
    globalhistory_data = list(global_history.find())
    if globalhistory_data:
        now = datetime.now()

        tdate = now.strftime("%d/%m/%Y")
        
        for daad in globalhistory_data:
            # print(daad)
            if daad["action"] < 0:
                if tdate == daad["date"]:
                    # print(daad)
                    add_amount_day = float(daad["amount"])
                    sum_amount_day = sum_amount_day + add_amount_day


        for daam in globalhistory_data:
            # print("Entered Loop")
            if daam["action"] < 0:
                # print("Found Action")
                if now.month == int(daam["date"].split('/')[1]):
                    # print("Enterted Sum area")
                    add_amount_month = float(daam["amount"])
                    sum_amount_month = sum_amount_month + add_amount_month
                 
        week_list = get_week_dates(tdate)

        for daaw in globalhistory_data:
            if daaw["action"] < 0:
                if daaw["date"] in week_list:
                    add_amount_week = float(daaw["amount"])
                    sum_amount_week = sum_amount_week + add_amount_week
                    
        
    # print(sum_amount_day)
    # print(sum_amount_week)
    # print(sum_amount_month)
    tdate2 = now.strftime("%Y-%m-%d")
    wn, month = get_current_week_of_month()
    wn_name = "Week"+str(wn)

    yearr = now.year
    update_monthly_income(month, yearr, sum_amount_month)
    upsert_income(tdate2, sum_amount_day)
    upsert_income_weekly(wn_name, month, sum_amount_week)

    update_earning = earnings.update_one({"Earning_ID":101},
                                         {
                                             "$set":{
                                                 "Daily_Income":sum_amount_day,
                                                 "Weekly_Income":sum_amount_week,
                                                 "Monthly_Income":sum_amount_month
                                             }
                                         })
    sum_amount_day = 0  
    sum_amount_week = 0
    sum_amount_month = 0





def graph_control():
    graph_control_daily()


def show_income_plot(dates, earnings):
    # Clear existing widgets (like old graphs) in frame_for_graph
    for widget in frame_for_graph.winfo_children():
        widget.destroy()

    # Convert string dates to datetime objects
    date_objs = [datetime.strptime(date, "%Y-%m-%d") for date in dates]

    # Create a Matplotlib figure
    fig = Figure(figsize=(8, 2), dpi=90)
    ax = fig.add_subplot(111)

    line, = ax.plot(date_objs, earnings, marker='o', color='blue', picker=5)
    ax.set_title("Daily Shop Income")
    ax.set_xlabel("Date")
    ax.set_ylabel("Earnings (₹)")
    ax.grid(True)
    fig.autofmt_xdate()

    def on_pick(event):
        ind = event.ind[0]
        label_for_graph.configure(text=f"Date: {dates[ind]}, Earnings: ₹{earnings[ind]}")

    fig.canvas.mpl_connect('pick_event', on_pick)

    canvas = FigureCanvasTkAgg(fig, master=frame_for_graph)
    canvas.draw()
    canvas.get_tk_widget().pack(side=TOP, fill=BOTH, expand=True)

    toolbar = NavigationToolbar2Tk(canvas, frame_for_graph)
    toolbar.update()
    toolbar.pack(side=TOP, fill=X)

def plot_weekly_income_chart(weeks_all, income_all):
    for widget in frame_for_graph_weekly.winfo_children():
        widget.destroy()

    def on_bar_click(event):
        # Get the axis where the click happened
        for i, bar in enumerate(bars):
            if bar.contains(event)[0]:
                week = weeks_all[i]
                income = income_all[i]
                # print(f"Clicked: {week} → Income: ${income}")
                label_for_graph.configure(text=f"{week}, Earnings: ₹{income}")
                break

    # Create figure and axis
    fig = Figure(figsize=(1, 1), dpi=90)
    ax = fig.add_subplot(111)
    bars = ax.barh(weeks_all, income_all, color='skyblue')
    ax.set_title("Weekly Income Overview")
    ax.set_xlabel("Income ($)")
    ax.set_ylabel("Weeks")
    ax.invert_yaxis()
    ax.grid(axis='x', linestyle='--', alpha=0.7)

    # Bind click event to the figure
    fig.canvas.mpl_connect('button_press_event', on_bar_click)

    # Embed plot in Tkinter
    canvas = FigureCanvasTkAgg(fig, master=frame_for_graph_weekly)
    canvas.draw()
    canvas.get_tk_widget().pack(fill='both', expand=True)

    # Add toolbar
    toolbar = NavigationToolbar2Tk(canvas, frame_for_graph_weekly)
    toolbar.update()
    toolbar.pack(side=BOTTOM, fill=X)

def plot_weekly_income_comparison(weeks, monthly_incomes):
    """
    Plots a line chart of weekly incomes for different months on a Tkinter window.
    Clicking on a point shows the week, month, and income.
    """
    for widget in frame_for_graph_weekly_c3.winfo_children():
        widget.destroy()

    fig = Figure(figsize=(6, 2), dpi=100)
    ax = fig.add_subplot(111)
    
    # Dictionary to store points for click lookup
    plotted_points = []

    for month, incomes in monthly_incomes.items():
        line, = ax.plot(weeks, incomes, marker='o', label=month, picker=5)  # `picker=5` enables click radius
        for i, income in enumerate(incomes):
            plotted_points.append({
                "week": weeks[i],
                "month": month,
                "income": income,
                "x": i,
                "y": income,
                "line": line
            })

    ax.set_title("Weekly Income Comparison")
    ax.set_xlabel("Weeks")
    ax.set_ylabel("Income ($)")
    ax.grid(True, linestyle='--', alpha=0.5)
    ax.legend()

    canvas = FigureCanvasTkAgg(fig, master=frame_for_graph_weekly_c3)
    canvas.draw()
    canvas.get_tk_widget().pack(fill='both', expand=True)

    toolbar = NavigationToolbar2Tk(canvas, frame_for_graph_weekly_c3)
    toolbar.update()
    toolbar.pack(side=BOTTOM, fill=X)

    # Define event handler
    def on_pick(event):
        mouse_event = event.mouseevent
        artist = event.artist
        ind = event.ind[0]

        for point in plotted_points:
            if point["line"] == artist and extract_week_index(point["week"]) == ind:
                messagebox.showinfo(
                    "Data Point Info",
                    f"Month: {point['month']}\nWeek: {point['week']}\nEarning: ${point['income']}"
                )
                break


    # Helper to map "Week 1" → 0, "Week 2" → 1, etc.
    def extract_week_index(week_str):
        return int(week_str.strip().replace("Week", "")) - 1

    # Connect the event
    fig.canvas.mpl_connect('pick_event', on_pick)

def plot_weekly_income_comparison_c6(weeks, monthly_incomes):
    """Plots weekly income comparison chart with interactive data point click."""
    for widget in frame_for_graph_weekly_c6.winfo_children():
        widget.destroy()

    fig = Figure(figsize=(8, 2), dpi=100)
    ax = fig.add_subplot(111)
    plotted_points = []

    for month, incomes in monthly_incomes.items():
        line, = ax.plot(weeks, incomes, marker='o', label=month, picker=5)
        for i, income in enumerate(incomes):
            plotted_points.append({
                "week": weeks[i],
                "month": month,
                "income": income,
                "line": line,
                "index": i
            })

    ax.set_title("Weekly Income Comparison (Last 6 Months)")
    ax.set_xlabel("Weeks")
    ax.set_ylabel("Income ($)")
    ax.grid(True, linestyle='--', alpha=0.5)
    ax.legend()

    canvas = FigureCanvasTkAgg(fig, master=frame_for_graph_weekly_c6)
    canvas.draw()
    canvas.get_tk_widget().pack(fill='both', expand=True)

    toolbar = NavigationToolbar2Tk(canvas, frame_for_graph_weekly_c6)
    toolbar.update()
    toolbar.pack(side=BOTTOM, fill=X)

    def on_pick(event):
        artist = event.artist
        ind = event.ind[0]
        for point in plotted_points:
            if point["line"] == artist and point["index"] == ind:
                messagebox.showinfo(
                    "Earning Info",
                    f"Month: {point['month']}\nWeek: {point['week']}\nEarning: ${point['income']}"
                )
                break

    fig.canvas.mpl_connect('pick_event', on_pick)

def plot_weekly_income_comparison_c12(weeks, monthly_incomes):
    """Plots a 12-month income comparison chart with clickable data points."""

    for widget in frame_for_graph_weekly_c12.winfo_children():
        widget.destroy()

    fig = Figure(figsize=(10, 2), dpi=100)
    ax = fig.add_subplot(111)
    plotted_points = []

    for month, incomes in monthly_incomes.items():
        line, = ax.plot(weeks, incomes, marker='o', label=month, picker=5)
        for i, income in enumerate(incomes):
            plotted_points.append({
                "week": weeks[i],
                "month": month,
                "income": income,
                "line": line,
                "index": i
            })

    ax.set_title("Weekly Income Comparison (Last 12 Months)")
    ax.set_xlabel("Weeks")
    ax.set_ylabel("Income ($)")
    ax.grid(True, linestyle='--', alpha=0.5)
    ax.legend(loc='upper left', bbox_to_anchor=(1, 1))  # Legend outside

    canvas = FigureCanvasTkAgg(fig, master=frame_for_graph_weekly_c12)
    canvas.draw()
    canvas.get_tk_widget().pack(fill='both', expand=True)

    toolbar = NavigationToolbar2Tk(canvas, frame_for_graph_weekly_c12)
    toolbar.update()
    toolbar.pack(side=BOTTOM, fill=X)

    def on_pick(event):
        artist = event.artist
        ind = event.ind[0]
        for point in plotted_points:
            if point["line"] == artist and point["index"] == ind:
                messagebox.showinfo(
                    "Earning Info",
                    f"Month: {point['month']}\nWeek: {point['week']}\nEarning: ${point['income']}"
                )
                break

    fig.canvas.mpl_connect('pick_event', on_pick)

def plot_monthly_income_chart(months, monthly_totals):
    """Plots total income per month for the last 12 months."""
    fig = Figure(figsize=(5, 3), dpi=100)
    ax = fig.add_subplot(111)

    bars = ax.bar(months, monthly_totals, color='skyblue', picker=True)

    ax.set_title("Monthly Income (Last 12 Months)")
    ax.set_xlabel("Month")
    ax.set_ylabel("Total Income ($)")
    ax.grid(axis='y', linestyle='--', alpha=0.7)

    canvas = FigureCanvasTkAgg(fig, master=frame_for_graph_monthly)
    canvas.draw()
    canvas.get_tk_widget().pack(fill='both', expand=True)

    toolbar = NavigationToolbar2Tk(canvas, frame_for_graph_monthly)
    toolbar.update()
    toolbar.pack(side=BOTTOM, fill=X)

    def on_pick(event):
        bar = event.artist
        for i, rect in enumerate(bars):
            if rect == bar:
                month = months[i]
                total = monthly_totals[i]
                messagebox.showinfo("Monthly Income", f"Month: {month}\nTotal Income: ${total:.2f}")
                break

    fig.canvas.mpl_connect('pick_event', on_pick)

def upsert_income(date, income):
    # Connect to the database
    current_dir = Path(__file__).resolve().parent if "__file__" in locals() else Path.cwd()
    conn = sqlite3.connect(current_dir/"Data"/"dailyEarnings.db")
    cursor = conn.cursor()

    # Check if the date exists
    cursor.execute("SELECT 1 FROM IncomeTable WHERE Dates = ?", (date,))
    result = cursor.fetchone()

    if result:
        # Update existing record
        cursor.execute("UPDATE IncomeTable SET Income = ? WHERE Dates = ?", (income, date))
        # print(f"Updated income for date {date}")
    else:
        # Insert new record
        cursor.execute("INSERT INTO IncomeTable (Dates, Income) VALUES (?, ?)", (date, income))
        # print(f"Inserted new income for date {date}")

    # Commit and close
    conn.commit()
    conn.close()


def upsert_income_weekly(week, month, income):
    current_dir = Path(__file__).resolve().parent if "__file__" in locals() else Path.cwd()

    conn = sqlite3.connect(current_dir/"Data"/"weeklyEarnings.db")
    cursor = conn.cursor()

    # Create the table if it doesn't exist
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS IncomeTable (
        Week TEXT,
        Month TEXT,
        Income REAL,
        PRIMARY KEY (Week, Month)
    )
    ''')

    # Function to insert or update income

    cursor.execute('''
    INSERT INTO IncomeTable (Week, Month, Income)
    VALUES (?, ?, ?)
    ON CONFLICT(Week, Month) DO UPDATE SET
        Income = excluded.Income
    ''', (week, month, income))
    conn.commit()
    conn.close()


check_current_choice = []

    
def graph_control_daily():
    cb3_var.set(0)
    cb6_var.set(0)
    cb12_var.set(0)
    compare_cb_label.place_forget()
    frame_for_graph_weekly_c3.pack_forget()
    frame_for_graph_weekly_c6.pack_forget()
    frame_for_graph_weekly_c12.pack_forget()
    cb_3m.place_forget()
    cb_6m.place_forget()
    cb_12m.place_forget()
    frame_for_graph_monthly.pack_forget()
    frame_for_graph_weekly.pack_forget()
    frame_for_graph.pack(fill = X, expand = True, side = 'bottom', padx = 10, pady = (20,0),anchor = 's')
    check_current_choice.insert(0, "graph_control_daily")

    current_dir = Path(__file__).resolve().parent if "__file__" in locals() else Path.cwd()

    conn = sqlite3.connect(current_dir/"Data"/"dailyEarnings.db")
    cursor = conn.cursor()

    # Fetch Dates and Income from the table
    cursor.execute("SELECT Dates, Income FROM IncomeTable")
    rows = cursor.fetchall()

    # Separate the results into two lists
    dates = [row[0] for row in rows]
    income = [row[1] for row in rows]

    # Close the connection
    conn.close()

    if len(dates) >= 7:
        show_income_plot(dates, income)
    else:
        messagebox.showinfo(f"Graph Activation ({7- len(dates)} days left)", "Don't have much data to show graph.")

def graph_control_weekly():
    cb3_var.set(0)
    cb6_var.set(0)
    cb12_var.set(0)
    compare_cb_label.place_forget()
    frame_for_graph_weekly_c3.pack_forget()
    frame_for_graph_weekly_c6.pack_forget()
    frame_for_graph_weekly_c12.pack_forget()
    cb_3m.place_forget()
    cb_6m.place_forget()
    cb_12m.place_forget()
    frame_for_graph.pack_forget()
    frame_for_graph_monthly.pack_forget()
    frame_for_graph_weekly.pack(fill = X, expand = True, side = 'bottom', padx = 10, pady = (20,0),anchor = 's')

    current_dir = Path(__file__).resolve().parent if "__file__" in locals() else Path.cwd()

    conn = sqlite3.connect(current_dir/"Data"/"weeklyEarnings.db")
    now = datetime.now()
    current_month_index = now.month
    all_months = [month_abbr[i] for i in range(1, 13)]  # ['Jan', 'Feb', ..., 'Dec']

    # Get the last 3 months including current
    recent_months = [all_months[(current_month_index - i - 1) % 12] for i in range(3)]

    # Connect to the database
    cursor = conn.cursor()

    # Prepare and execute the query
    query = f'''
        SELECT Week, Month, Income
        FROM IncomeTable
        WHERE Month IN ({','.join(['?'] * len(recent_months))})
    '''
    cursor.execute(query, recent_months)
    rows = cursor.fetchall()

    # Store results
    Week = []
    Month = []
    WeekMonth = []
    Income = []

    for week, month, income in rows:
        Week.append(week)
        Month.append(month)
        WeekMonth.append(f"{week} ({month})")
        Income.append(income)

    conn.close()

    if len(WeekMonth) >= 4:
        plot_weekly_income_chart(WeekMonth, Income)
        compare_cb_label.place(x = 690, y = 330 )
        cb_3m.place(x = 810, y = 330)
        cb_6m.place(x = 910, y = 330)
        cb_12m.place(x = 1010, y = 330)
    else:
        try:
            messagebox.showinfo(f"Graph Activation ({4- len(week)} weeks left)", "Don't have much data to show graph.")
        except UnboundLocalError:
            messagebox.showinfo(f"Graph Activation", "Don't have much data to show graph.")


def graph_control_monthly():
    cb3_var.set(0)
    cb6_var.set(0)
    cb12_var.set(0)
    compare_cb_label.place_forget()
    frame_for_graph_weekly_c3.pack_forget()
    frame_for_graph_weekly_c6.pack_forget()
    frame_for_graph_weekly_c12.pack_forget()
    cb_3m.place_forget()
    cb_6m.place_forget()
    cb_12m.place_forget()
    frame_for_graph.pack_forget()
    frame_for_graph_weekly.pack_forget()
    frame_for_graph_monthly.pack(fill = X, expand = True, side = 'bottom', padx = 10, pady = (20,0),anchor = 's')
    now = datetime.now()
    year = now.year
    get_monthly_income_by_year(year)

def handle_graph_cb():
    if cb3_var.get():
        cb6_var.set(0)
        cb12_var.set(0)
        frame_for_graph.pack_forget()
        frame_for_graph_monthly.pack_forget()
        frame_for_graph_weekly.pack_forget()
        frame_for_graph_weekly_c6.pack_forget()
        frame_for_graph_weekly_c12.pack_forget()
        frame_for_graph_weekly_c3.pack(fill = X, expand = True, side = 'bottom', padx = 10, pady = (20,0),anchor = 's')

        current_dir = Path(__file__).resolve().parent if "__file__" in locals() else Path.cwd()

        conn = sqlite3.connect(current_dir/"Data"/"weeklyEarnings.db")
        cursor = conn.cursor()

        # Get current and previous two months in abbreviated form
        now = datetime.now()
        month_indices = [(now.month - i - 1) % 12 + 1 for i in range(3)]  # e.g., [6, 5, 4]
        month_names = [month_abbr[i] for i in reversed(month_indices)]   # e.g., ['Apr', 'May', 'Jun']

        # Query data for these months
        placeholders = ','.join('?' for _ in month_names)
        query = f"SELECT Week, Month, Income FROM IncomeTable WHERE Month IN ({placeholders})"
        cursor.execute(query, month_names)

        # Prepare containers
        weeks_set = set()
        monthly_incomes = {month: [] for month in month_names}
        temp_data = {month: {} for month in month_names}

        # Extract numeric part from 'Week1' / 'Week 1'
        def extract_week_number(week_str):
            match = re.search(r'\d+', week_str)
            return int(match.group()) if match else 0

        # Load DB results
        for week, month, income in cursor.fetchall():
            weeks_set.add(week)
            temp_data[month][week] = income

        # Sort week labels like 'Week1', 'Week2', etc.
        weeks = sorted(list(weeks_set), key=extract_week_number)

        # Construct final monthly incomes ordered by weeks
        for month in month_names:
            monthly_incomes[month] = [temp_data[month].get(week, 0) for week in weeks]

        conn.close()

        plot_weekly_income_comparison(weeks, monthly_incomes)

    elif cb6_var.get():
        cb3_var.set(0)
        cb12_var.set(0)
        frame_for_graph.pack_forget()
        frame_for_graph_monthly.pack_forget()
        frame_for_graph_weekly.pack_forget()
        frame_for_graph_weekly_c3.pack_forget()
        frame_for_graph_weekly_c12.pack_forget()
        frame_for_graph_weekly_c6.pack(fill = X, expand = True, side = 'bottom', padx = 10, pady = (20,0),anchor = 's')

        current_dir = Path(__file__).resolve().parent if "__file__" in locals() else Path.cwd()

        conn = sqlite3.connect(current_dir/"Data"/"weeklyEarnings.db")
        cursor = conn.cursor()

        now = datetime.now()
        # Get 6 most recent months (including current), as 3-letter abbreviations
        month_indices = [(now.month - i - 1) % 12 + 1 for i in range(6)]
        month_names = [month_abbr[i] for i in reversed(month_indices)]  # e.g., ['Jan', 'Feb', ..., 'Jun']

        # Fetch data for these months
        placeholders = ','.join('?' for _ in month_names)
        query = f"SELECT Week, Month, Income FROM IncomeTable WHERE Month IN ({placeholders})"
        cursor.execute(query, month_names)

        # Prepare data structures
        weeks_set = set()
        monthly_incomes = {month: {} for month in month_names}

        for week, month, income in cursor.fetchall():
            weeks_set.add(week)
            monthly_incomes[month][week] = income

        # Sort week labels like 'Week1', 'Week2', etc.
        def extract_week_number(week_str):
            match = re.search(r'\d+', week_str)
            return int(match.group()) if match else 0

        sorted_weeks = sorted(weeks_set, key=extract_week_number)

        # Fill missing data with 0
        monthly_incomes_complete = {
            month: [monthly_incomes[month].get(week, 0) for week in sorted_weeks]
            for month in month_names
        }

        conn.close()

        plot_weekly_income_comparison_c6(sorted_weeks, monthly_incomes_complete)

    elif cb12_var.get():
        cb6_var.set(0)
        cb3_var.set(0)
        frame_for_graph.pack_forget()
        frame_for_graph_monthly.pack_forget()
        frame_for_graph_weekly.pack_forget()
        frame_for_graph_weekly_c3.pack_forget()
        frame_for_graph_weekly_c6.pack_forget()
        frame_for_graph_weekly_c12.pack(fill = X, expand = True, side = 'bottom', padx = 10, pady = (20,0),anchor = 's')

        current_dir = Path(__file__).resolve().parent if "__file__" in locals() else Path.cwd()

        conn = sqlite3.connect(current_dir/"Data"/"weeklyEarnings.db")
        cursor = conn.cursor()

        now = datetime.now()
        # Get past 12 months as 3-letter abbreviations (e.g., 'Jul', 'Aug', ...)
        month_indices = [(now.month - i - 1) % 12 + 1 for i in range(12)]
        month_names = [month_abbr[i] for i in reversed(month_indices)]

        # Query database
        placeholders = ','.join('?' for _ in month_names)
        query = f"SELECT Week, Month, Income FROM IncomeTable WHERE Month IN ({placeholders})"
        cursor.execute(query, month_names)

        # Prepare containers
        weeks_set = set()
        monthly_incomes = {month: {} for month in month_names}

        for week, month, income in cursor.fetchall():
            weeks_set.add(week)
            monthly_incomes[month][week] = income

        # Sort weeks like 'Week1', 'Week2', etc.
        def extract_week_number(week_str):
            match = re.search(r'\d+', week_str)
            return int(match.group()) if match else 0

        sorted_weeks = sorted(weeks_set, key=extract_week_number)

        # Complete structure: fill missing values with 0
        monthly_incomes_complete = {
            month: [monthly_incomes[month].get(week, 0) for week in sorted_weeks]
            for month in month_names
        }

        conn.close()
    
        plot_weekly_income_comparison_c12(sorted_weeks, monthly_incomes_complete)
    else:
        frame_for_graph_weekly_c3.pack_forget()
        frame_for_graph_weekly_c6.pack_forget()
        frame_for_graph_weekly_c12.pack_forget()

def get_monthly_income_by_year(year):
    cwd = os.getcwd()
    current_path = os.path.join(cwd, "Data", "monthly_income_data.db")
    conn = sqlite3.connect(current_path)
    cursor = conn.cursor()

    # Retrieve data for the specified year, ordered by month if needed
    cursor.execute('''
        SELECT month, income
        FROM income
        WHERE year = ?
    ''', (year,))

    results = cursor.fetchall()

    # Separate into two lists
    month_list = [row[0][0:3] for row in results]
    income_list = [row[1] for row in results]

    # Close the connection
    conn.close()

    plot_monthly_income_chart(month_list, income_list)



def update_monthly_income(month, year, earning):
    cwd = os.getcwd()
    current_path = os.path.join(cwd, "Data", "monthly_income_data.db")
    conn = sqlite3.connect(current_path)
    cursor = conn.cursor()
    
    # Check if a record for this month and year already exists
    cursor.execute('''
        SELECT id FROM income WHERE month = ? AND year = ?
    ''', (month, year))
    
    result = cursor.fetchone()
    
    if result:
        # Record exists, update it
        cursor.execute('''
            UPDATE income SET income = ? WHERE month = ? AND year = ?
        ''', (earning, month, year))
    else:
        # Record doesn't exist, insert new
        cursor.execute('''
            INSERT INTO income (month, year, income) VALUES (?, ?, ?)
        ''', (month, year, earning))
    
    conn.commit()
    conn.close()

if check_internet():
    current_dir = Path(__file__).resolve().parent if "__file__" in locals() else Path.cwd()
    envars_db = current_dir/ ".env" # create .env file in current folder

    load_dotenv(envars_db)

    client = pymongo.MongoClient(os.getenv('URL_Mongo'))

    login_signup_database = client['Login-Signup']
    client_info = login_signup_database.clien_infos
    win=Tk()
    win.geometry("1400x770+50+0")
    win.title("Bazaro")
    # win.iconbitmap('logo.ico')

    # =======================<<<<<<<<<<<< LOGIN INTERFACE FROM HERE >>>>>>>>>>>>>>>=============================

    contentframe=Frame(win,height=770,width=1400,bg='red')
    contentframe.propagate(False)
    contentframe.pack()


    login_canvas=Canvas(contentframe,height=770,width=1400,bg='black',bd=0,highlightthickness=0, relief='ridge')
    login_canvas.propagate(False)
    login_canvas.pack()

    cwd = os.getcwd()

    imagepath=cwd+"\\uiux\\login 2.png"
    openphoto=Image.open(imagepath).resize((1400,770))
    bgimage=ImageTk.PhotoImage(openphoto)
    login_canvas.create_image(700,385, image=bgimage)

    username = StringVar()

    username_entry = ctk.CTkEntry(login_canvas, height=40,width=330, fg_color='#f2f3f4', text_color='black',
                                border_color='#FF8A00', bg_color='white', corner_radius=20, font=('Poppins', 22),
                                textvariable= username)
    username_entry.place(x = 600, y = 435)

    password = StringVar()

    password_entry = ctk.CTkEntry(login_canvas, height=40,width=330, fg_color='#f2f3f4', text_color='black',
                                border_color='#FF8A00', bg_color='white', corner_radius=20, font=('Poppins', 22),
                                textvariable= password)
    password_entry.place(x = 600, y = 565)

    pass_forget_button = ctk.CTkButton(login_canvas, text="Forgot password?", font=('arial', 12), height=1, width = 20, cursor='hand2',
                                fg_color='white', bg_color='white', text_color='#FF8A00', hover_color='white', command = new_password)
    pass_forget_button.place(x = 400, y = 620)

    new_member_button = ctk.CTkButton(login_canvas, text="New to Bazaro? Sign up!", font=('arial', 12), height=1, width = 20, cursor='hand2',
                                fg_color='white', bg_color='white', text_color='blue', hover_color='white', command=sign_up)
    new_member_button.place(x = 860, y = 620)


    continue_button = ctk.CTkButton(login_canvas, text="Continue", font=('arial', 26, 'bold'), height=45, width = 40, cursor='hand2',
                                fg_color='orange', bg_color='white', text_color='white', hover_color='black', corner_radius=10,
                                command=try_login)
    continue_button.place(x = 635, y = 660)

    # =======================<<<<<<<<<<<< forget password INTERFACE FROM HERE >>>>>>>>>>>>>>>=============================

    pass_reset_frame=Frame(win,height=770,width=1400,bg='red')
    pass_reset_frame.propagate(False)
    # pass_reset_frame.pack()


    reset_canvas=Canvas(pass_reset_frame,height=770,width=1400,bg='black',bd=0,highlightthickness=0, relief='ridge')
    reset_canvas.propagate(False)
    reset_canvas.pack()
    
    imagepath_1=cwd+"\\uiux\\forget_pass.png"
    openphoto_1=Image.open(imagepath_1).resize((1400,770))
    bgimage_1=ImageTk.PhotoImage(openphoto_1)
    reset_canvas.create_image(700,385, image=bgimage_1)
    
    resetEmailotp = StringVar()
    reset_email_otp = ctk.CTkEntry(reset_canvas, height=40,width=330, fg_color='#f2f3f4', text_color='black',
                                border_color='#FF8A00', bg_color='white', corner_radius=20, font=('Poppins', 22),
                                textvariable= resetEmailotp)
    reset_email_otp.place(x = 900, y = 400)
    
    reset_verify_button = ctk.CTkButton(reset_canvas, text="Verify", font=('Poppins', 26, 'bold'), height=45, width = 40, cursor='hand2',
                                fg_color='orange', bg_color='white', text_color='white', hover_color='black',
                                corner_radius=10, command = reset_verify_otp)
    reset_verify_button.place(x = 900, y = 480)

    resend_button = ctk.CTkButton(reset_canvas, text="Resend", font=('Poppins', 26, 'bold'), height=45, width = 40, cursor='hand2',
                                fg_color='orange', bg_color='white', text_color='white', hover_color='black',
                                corner_radius=10, command= reset_resend_otp)
    resend_button.place(x = 1020, y = 480)
    
    # =======================<<<<<<<<<<<< new password INTERFACE FROM HERE >>>>>>>>>>>>>>>=============================

    new_pass_frame=Frame(win,height=770,width=1400,bg='red')
    new_pass_frame.propagate(False)

    new_pass_canvas=Canvas(new_pass_frame,height=770,width=1400,bg='black',bd=0,highlightthickness=0, relief='ridge')
    new_pass_canvas.propagate(False)
    new_pass_canvas.pack()
    
    imagepath_2=cwd+"\\uiux\\new_pass.png"
    openphoto_2=Image.open(imagepath_2).resize((1400,770))
    bgimage_2=ImageTk.PhotoImage(openphoto_2)
    new_pass_canvas.create_image(700,385, image=bgimage_2)
    
    resetpass = StringVar()
    reset_password_entry = ctk.CTkEntry(new_pass_canvas, height=40,width=330, fg_color='#f2f3f4', text_color='black',
                                border_color='#FF8A00', bg_color='white', corner_radius=20, font=('Poppins', 22),
                                textvariable= resetpass)
    reset_password_entry.place(x = 970, y = 400)
    
    reset_password_button = ctk.CTkButton(new_pass_canvas, text="Reset", font=('Poppins', 26, 'bold'), height=45, width = 40, cursor='hand2',
                                fg_color='orange', bg_color='white', text_color='white', hover_color='black',
                                corner_radius=10, command = reset_password)
    reset_password_button.place(x = 950, y = 480)
    
    # =======================<<<<<<<<<<<< SIGNUP INTERFACE FROM HERE >>>>>>>>>>>>>>>=============================

    sign_up_frame=Frame(win,height=770,width=1400,bg='blue')
    sign_up_frame.propagate(False)
    # sign_up_frame.pack()

    sign_up_canvas=Canvas(sign_up_frame,height=770,width=1400,bg='blue',bd=0,highlightthickness=0, relief='ridge')
    sign_up_canvas.propagate(False)
    sign_up_canvas.pack()

    cwd = os.getcwd()

    imagepath1=cwd+"\\uiux\\sign up 2.png"
    openphoto1=Image.open(imagepath1).resize((1400,770))
    bgimage1=ImageTk.PhotoImage(openphoto1)
    sign_up_canvas.create_image(700,385, image=bgimage1)


    username_signup = StringVar()

    username_entry_signup = ctk.CTkEntry(sign_up_frame, height=40,width=330, fg_color='#f2f3f4', text_color='black',
                                border_color='#FF8A00', bg_color='white', corner_radius=20, font=('Poppins', 22),
                                textvariable= username_signup)
    username_entry_signup.place(x = 900, y = 145)

    email_signup = StringVar()

    email_entry_signup = ctk.CTkEntry(sign_up_frame, height=40,width=330, fg_color='#f2f3f4', text_color='black',
                                border_color='#FF8A00', bg_color='white', corner_radius=20, font=('Poppins', 22),
                                textvariable= email_signup)
    email_entry_signup.place(x = 900, y = 225)

    phone_signup = StringVar()

    mobile_entry_signup = ctk.CTkEntry(sign_up_frame, height=40,width=330, fg_color='#f2f3f4', text_color='black',
                                border_color='#FF8A00', bg_color='white', corner_radius=20, font=('Poppins', 22),
                                textvariable= phone_signup)
    mobile_entry_signup.place(x = 900, y = 315)

    password_signup = StringVar()

    password_entry_signup = ctk.CTkEntry(sign_up_frame, height=40,width=330, fg_color='#f2f3f4', text_color='black',
                                border_color='#FF8A00', bg_color='white', corner_radius=20, font=('Poppins', 22),
                                textvariable= password_signup)
    password_entry_signup.place(x = 900, y = 405)


    letsgo_button = ctk.CTkButton(sign_up_frame, text="Lets go!", font=('arial', 26, 'bold'), height=45, width = 40, cursor='hand2',
                                fg_color='orange', bg_color='white', text_color='white', hover_color='black', corner_radius=10,
                                command = try_signup)
    letsgo_button.place(x = 950, y = 500)

    # =======================<<<<<<<<<<<< Verification INTERFACE FROM HERE >>>>>>>>>>>>>>>=============================

    verification_frame=Frame(win,height=770,width=1400,bg='green')
    verification_frame.propagate(False)
    # sign_up_frame.pack()

    verification_canvas=Canvas(verification_frame,height=770,width=1400,bg='blue',bd=0,highlightthickness=0, relief='ridge')
    verification_canvas.propagate(False)
    verification_canvas.pack()

    imagepath4=cwd+"\\uiux\\verification2.png"
    openphoto4=Image.open(imagepath4).resize((1400,770))
    bgimage4=ImageTk.PhotoImage(openphoto4)
    verification_canvas.create_image(700,385, image=bgimage4)

    email_otp = StringVar()

    email_otp_entry = ctk.CTkEntry(verification_frame, height=40,width=330, fg_color='#f2f3f4', text_color='black',
                                border_color='#FF8A00', bg_color='white', corner_radius=20, font=('Poppins', 22),
                                textvariable= email_otp)
    email_otp_entry.place(x = 910, y = 300)

    phone_otp = StringVar()

    email_otp_entry = ctk.CTkEntry(verification_frame, height=40,width=330, fg_color='#f2f3f4', text_color='black',
                                border_color='#FF8A00', bg_color='white', corner_radius=20, font=('Poppins', 22),
                                textvariable= phone_otp)
    email_otp_entry.place(x = 910, y = 430)
    
    back_button = ctk.CTkButton(verification_frame, text="Back", font=('Poppins', 26, 'bold'), height=45, width = 40, cursor='hand2',
                                fg_color='orange', bg_color='white', text_color='white', hover_color='black',
                                corner_radius=10, command = get_back)
    back_button.place(x = 840, y = 510)

    verify_button = ctk.CTkButton(verification_frame, text="Verify", font=('Poppins', 26, 'bold'), height=45, width = 40, cursor='hand2',
                                fg_color='orange', bg_color='white', text_color='white', hover_color='black',
                                corner_radius=10, command = verify_otp)
    verify_button.place(x = 950, y = 510)

    resend_button = ctk.CTkButton(verification_frame, text="Resend", font=('Poppins', 26, 'bold'), height=45, width = 40, cursor='hand2',
                                fg_color='orange', bg_color='white', text_color='white', hover_color='black',
                                corner_radius=10, command= resend_otp)
    resend_button.place(x = 1070, y = 510)
    
    # =======================<<<<<<<<<<<< Dashboard INTERFACE FROM HERE >>>>>>>>>>>>>>>=============================

    file_path = 'shopname.txt'
    shop_name = ''
    try:
        with open(file_path, 'r') as file:
            content = file.read().strip().split()
            # shop_name_for_bill = file.read().strip()
            for parts in content:
                shop_name = shop_name+'-'+parts
            
    except FileNotFoundError:
        messagebox.showerror('File Error',f"The file '{file_path}' was not found.")
    except IOError:
        messagebox.showerror('Read Error',f"An error occurred while reading the file '{file_path}'.")

    r_shop_name = shop_name[1:]
    
    shop_database = client[r_shop_name]
    inventory = shop_database.stock_inventory
    global_history = shop_database.Global_History
    earnings = shop_database.Earnings
    
    db_file_er = current_dir/"Data"/"decide.db"

    # Connect to the SQLite database
    conn_ern = sqlite3.connect(db_file_er)
    cursor_ern = conn_ern.cursor()
    
    cursor_ern.execute('''
        SELECT decidee FROM mongoEarnings WHERE EID = ?
    ''', (104,))

    result_ern = cursor_ern.fetchone()
    
    if result_ern[0] == 1:
        data_ern = {
            "Earning_ID": 101,
            "Daily_Income": 0,
            "Weekly_Income": 0,
            "Monthly_Income": 0
        }

        # Insert the data
        insert_result = earnings.insert_one(data_ern)
        print("INSERTED")
        cursor_ern.execute('''
            UPDATE mongoEarnings
            SET decidee = ?
            WHERE EID = ?
        ''', (0, 104))  # You can change 1 to any other value

        # Commit the changes and close the connection
        conn_ern.commit()
    conn_ern.close()
        
        
    customer_care_database = client['Customer-Care']
    cc_database = customer_care_database.cc_messages
    
    
    dashboard_frame=Frame(win,height=770,width=1400,bg='red')
    dashboard_frame.propagate(False)
    # dashboard_frame.pack()
    
    menu_frame=Frame(dashboard_frame,height=770,width=250,bg='red')
    menu_frame.propagate(False)
    menu_frame.pack(side='left')

    menu_canvas=Canvas(menu_frame,height=770,width=250,bg='blue',bd=0,highlightthickness=0, relief='ridge')
    menu_canvas.propagate(False)
    menu_canvas.pack()
    
    imagepath2=cwd+"\\uiux\\sidebar2.png"
    openphoto2=Image.open(imagepath2).resize((250,770))
    bgimage2=ImageTk.PhotoImage(openphoto2)
    menu_canvas.create_image(125,385, image=bgimage2)

    dashboard_button = ctk.CTkButton(menu_frame, text="DASHBOARD", font=('Poppins', 20), height=45, width = 170, cursor='hand2',
                                fg_color='#4B54F8', bg_color='#4B54F8', text_color='white', hover_color='#0713FF',
                                corner_radius=10, anchor='w', command= dashboardFunction )
    dashboard_button.place(x = 60, y = 100)
    
    inventory_button = ctk.CTkButton(menu_frame, text="INVENTORY", font=('Poppins', 20), height=45, width = 170, cursor='hand2',
                                fg_color='#4B54F8', bg_color='#4B54F8', text_color='white', hover_color='#0713FF',
                                corner_radius=10, anchor='w', command= inventoryFunction )
    inventory_button.place(x = 60, y = 175)
    
    alert_button = ctk.CTkButton(menu_frame, text="ALERTS", font=('Poppins', 20), height=45, width = 170, cursor='hand2',
                                fg_color='#4B54F8', bg_color='#4B54F8', text_color='white', hover_color='#0713FF',
                                corner_radius=10, anchor='w', command =alertFunction )
    alert_button.place(x = 60, y = 250)
    
    billing_button = ctk.CTkButton(menu_frame, text="Billings", font=('Poppins', 20), height=45, width = 170, cursor='hand2',
                                fg_color='#4B54F8', bg_color='#4B54F8', text_color='white', hover_color='#0713FF',
                                corner_radius=10, anchor='w', command = billingFunction )
    billing_button.place(x = 60, y = 325)
    
    suplier_button = ctk.CTkButton(menu_frame, text="SUPPLIER", font=('Poppins', 20), height=45, width = 170, cursor='hand2',
                                fg_color='#4B54F8', bg_color='#4B54F8', text_color='white', hover_color='#0713FF',
                                corner_radius=10, anchor='w', command= supplierFunction )
    suplier_button.place(x = 60, y = 395)
    
    history_button = ctk.CTkButton(menu_frame, text="HISTORY", font=('Poppins', 20), height=45, width = 170, cursor='hand2',
                                fg_color='#4B54F8', bg_color='#4B54F8', text_color='white', hover_color='#0713FF',
                                corner_radius=10, anchor='w', command= historyFunction )
    history_button.place(x = 60, y = 475)
    
    
    settings_button = ctk.CTkButton(menu_frame, text="Settings", font=('Poppins', 20), height=45, width = 170, cursor='hand2',
                                fg_color='#4B54F8', bg_color='#4B54F8', text_color='white', hover_color='#0713FF',
                                corner_radius=10, anchor='w', command= settingFunction )
    settings_button.place(x = 60, y = 560)
    
    cc_button = ctk.CTkButton(menu_frame, text="Customer Care", font=('Poppins', 20), height=45, width = 170, cursor='hand2',
                                fg_color='#4B54F8', bg_color='#4B54F8', text_color='white', hover_color='#0713FF',
                                corner_radius=10, anchor='w', command= ccFunction )
    cc_button.place(x = 60, y = 640)
    
    exit_button = ctk.CTkButton(menu_frame, text="Exit", font=('Poppins', 20), height=45, width = 170, cursor='hand2',
                                fg_color='#4B54F8', bg_color='#4B54F8', text_color='white', hover_color='#0713FF',
                                corner_radius=10, anchor='w' )
    exit_button.place(x = 60, y = 715)          
    
    
    # =======================<<<<<<<<<<<< Dashboard Display INTERFACE FROM HERE >>>>>>>>>>>>>>>=============================

    dashboard_display=Frame(dashboard_frame,height=770,width=1150,bg='blue')
    dashboard_display.propagate(False)
    dashboard_display.pack(side='left')
    
    dashboard_canvas=Canvas(dashboard_display,height=770,width=1150,bg='blue',bd=0,highlightthickness=0, relief='ridge')
    dashboard_canvas.propagate(False)
    dashboard_canvas.pack()

    frame_for_graph =Frame(dashboard_canvas, height = 350, bg ='blue')
    frame_for_graph.propagate(False)

    frame_for_graph_weekly = Frame(dashboard_canvas, height = 350, bg ='blue')
    frame_for_graph_weekly.propagate(False)

    frame_for_graph_weekly_c3 = Frame(dashboard_canvas, height = 350, bg ='blue')
    frame_for_graph_weekly_c3.propagate(False)

    frame_for_graph_weekly_c6 = Frame(dashboard_canvas, height = 350, bg ='blue')
    frame_for_graph_weekly_c6.propagate(False)

    frame_for_graph_weekly_c12 = Frame(dashboard_canvas, height = 350, bg ='blue')
    frame_for_graph_weekly_c12.propagate(False)

    frame_for_graph_monthly = Frame(dashboard_canvas, height = 350, bg ='blue')
    frame_for_graph_monthly.propagate(False)


    label_for_graph = ctk.CTkLabel(dashboard_canvas, text="Graph Info", font=('poppins', 15, 'bold'), fg_color="#58d68d",
                                   bg_color="white", text_color="black", height=20, width=300, corner_radius=20)
    label_for_graph.pack(side = 'bottom', anchor = 's', pady = 10)

    compare_cb_label = ctk.CTkLabel(dashboard_canvas, text="Compare Past:", font=('poppins', 15, 'bold'), fg_color="White", bg_color="white",
                                    text_color="black")

    cb3_var = IntVar()
    cb_3m = ctk.CTkCheckBox(dashboard_display, text="3 months", fg_color="white", bg_color="white", text_color="black",
                            checkmark_color='green', border_color="black", variable=cb3_var, command= handle_graph_cb)
 

    cb6_var = IntVar()
    cb_6m = ctk.CTkCheckBox(dashboard_display, text="6 months", fg_color="white", bg_color="white", text_color="black",
                            checkmark_color='green', border_color="black", variable=cb6_var, command= handle_graph_cb)


    cb12_var = IntVar()
    cb_12m = ctk.CTkCheckBox(dashboard_display, text="12 months", fg_color="white", bg_color="white", text_color="black",
                            checkmark_color='green', border_color="black", variable=cb12_var, command= handle_graph_cb)





    imagepath3=cwd+"\\uiux\\dashboard3.png"
    openphoto3=Image.open(imagepath3).resize((1150,770))
    bgimage3=ImageTk.PhotoImage(openphoto3)
    dashboard_canvas.create_image(575,385, image=bgimage3)
    
    
    username_label = Label(dashboard_display, text='', font=('Poppins', 20), fg="black", bg = 'white')
    username_label.place(x = 100, y = 25)
    
    day_label = Label(dashboard_display, text='Tuesday', font = ('Poppins', 15), fg="#6F6F6F", bg='white')
    day_label.place(x = 1000, y = 10 )

    date_time_label =  Label(dashboard_display, text='20-05-2024 19:34', font = ('Poppins', 15), fg=  "black", bg='white')
    date_time_label.place(x = 915, y = 35)
    
    # progress_var = ctk.BooleanVar()
    daily_income_profit = ctk.CTkLabel(dashboard_display, text='Rs.400', font = ('poppins', 20, 'bold'), text_color="green",
                                       fg_color="white")
    daily_income_profit.place( x = 350, y = 160)
    
    daily_progress = ctk.CTkProgressBar(dashboard_display, orientation="horizontal", bg_color='white',
                                        width=300, progress_color="#4B54F8") # Initial set (0-1 scale)
    daily_progress.place( x = 35, y = 200)
    
    daily_progress_label = Label(dashboard_display, text='',font = ('Poppins', 12), fg="#6F6F6F", bg='white' )
    daily_progress_label.place(x = 340, y = 190)
    
    
    weekly_income_profit = ctk.CTkLabel(dashboard_display, text='Rs.5400', font = ('poppins', 20, 'bold'), text_color="green",
                                       fg_color="white")
    weekly_income_profit.place( x = 350, y = 225)
    
    weekly_progress = ctk.CTkProgressBar(dashboard_display, orientation="horizontal", bg_color='white',
                                        width=300, progress_color="#4B54F8")
    
    weekly_progress.place( x = 35, y = 270)
    
    weekly_progress_label = Label(dashboard_display, text='',font = ('Poppins', 12), fg="#6F6F6F", bg='white' )
    weekly_progress_label.place(x = 345, y = 260)
    
    monthly_income_profit = ctk.CTkLabel(dashboard_display, text='Rs.55400', font = ('poppins', 20, 'bold'), text_color="green",
                                       fg_color="white")
    monthly_income_profit.place( x = 350, y = 295)
    
    monthly_progress = ctk.CTkProgressBar(dashboard_display, orientation="horizontal", bg_color='white',
                                        width=300, progress_color="#4B54F8")
        
    monthly_progress.place( x = 35, y = 340 )

    
    monthly_progress_label = Label(dashboard_display, text='',font = ('Poppins', 12), fg="#6F6F6F", bg='white' )
    monthly_progress_label.place(x = 340, y = 330)

    graph_refresh_button = ctk.CTkButton(dashboard_display, text='Refresh Graph', font=('poppins', 15, 'bold'), fg_color="#4B54F8",
                                         bg_color="white", text_color='white', hover_color="#8B90F6",
                                         command = graph_control)
    graph_refresh_button.place(x = 950, y = 125)


    daily_income_graph_button = ctk.CTkButton(dashboard_display, text='Daily Income Graph', font=('poppins', 15, 'bold'), fg_color="white",
                                         bg_color="white", text_color='black', hover_color="#8B90F6", border_width=2, border_color="black",
                                         command = graph_control_daily)
    daily_income_graph_button.place(x = 690, y = 160)

    weekly_income_graph_button = ctk.CTkButton(dashboard_display, text='Weekly Income Graph', font=('poppins', 15, 'bold'), fg_color="white",
                                         bg_color="white", text_color='black', hover_color="#8B90F6", border_width=2, border_color="black",
                                         command = graph_control_weekly)
    weekly_income_graph_button.place(x = 900, y = 160)

    monthly_income_graph_button = ctk.CTkButton(dashboard_display, text='Monthly Income Graph', font=('poppins', 15, 'bold'), fg_color="white",
                                         bg_color="white", text_color='black', hover_color="#8B90F6", border_width=2, border_color="black",
                                         command = graph_control_monthly)
    monthly_income_graph_button.place(x = 785, y = 200)
    
    # =======================<<<<<<<<<<<< Inventory Display INTERFACE FROM HERE >>>>>>>>>>>>>>>=============================

    inventory_display=Frame(dashboard_frame,height=770,width=1150,bg='blue')
    inventory_display.propagate(False)
    # inventory_display.pack(side='left')
    
    inventory_canvas=Canvas(inventory_display,height=770,width=1150,bg='blue',bd=0,highlightthickness=0, relief='ridge')
    inventory_canvas.propagate(False)
    inventory_canvas.pack()


    imagepath10=cwd+"\\uiux\\inventory.png"
    openphoto10=Image.open(imagepath10).resize((1150,770))
    bgimage10=ImageTk.PhotoImage(openphoto10)
    inventory_canvas.create_image(575,385, image=bgimage10)
    
    add_stock_button = ctk.CTkButton(inventory_display, text="Add Stocks", font=('Poppins', 25,'bold'), height=90, width = 190, cursor='hand2',
                                fg_color='white', bg_color='white', text_color='Black', hover_color='white',
                                corner_radius=10, command= stock_add)
    add_stock_button.place(x = 719, y = 50)
    
    update_product_button = ctk.CTkButton(inventory_display, text="Update Stocks", font=('Poppins', 25,'bold'), height=90, width = 188, cursor='hand2',
                                fg_color='white', bg_color='white', text_color='Black', hover_color='white',
                                corner_radius=10, command = update_stock)
    update_product_button.place(x = 817, y = 195)
    
    stock_analytics = ctk.CTkButton(inventory_display, text="Stock Analytics", font=('Poppins', 22,'bold'), height=90, width = 190, cursor='hand2',
                                fg_color='white', bg_color='white', text_color='Black', hover_color='white',
                                corner_radius=10, command=analyse_stock)
    stock_analytics.place(x = 917, y = 340)
    
    check_stock_button = ctk.CTkButton(inventory_display, text="Future Updates", font=('Poppins', 22,'bold'), height=100, width = 190, cursor='hand2',
                                fg_color='white', bg_color='white', text_color='Black', hover_color='white',
                                corner_radius=10)
    check_stock_button.place(x = 817, y = 480)
    
    most_sold_button = ctk.CTkButton(inventory_display, text="Future Updates", font=('Poppins', 22,'bold'), height=90, width = 190, cursor='hand2',
                                fg_color='white', bg_color='white', text_color='Black', hover_color='white',
                                corner_radius=10)
    most_sold_button.place(x = 719, y = 630)
    
    # =======================<<<<<<<<<<<< Add Stock Display INTERFACE FROM HERE >>>>>>>>>>>>>>>=============================

    add_stock_frame=Frame(dashboard_frame,height=770,width=1150,bg='blue')
    add_stock_frame.propagate(False)
    # inventory_display.pack(side='left')
    
    add_stock_canvas=Canvas(add_stock_frame,height=770,width=1150,bg='blue',bd=0,highlightthickness=0, relief='ridge')
    add_stock_canvas.propagate(False)
    add_stock_canvas.pack()


    imagepath11=cwd+"\\uiux\\add_stock.png"
    openphoto11=Image.open(imagepath11).resize((1150,770))
    bgimage11=ImageTk.PhotoImage(openphoto11)
    add_stock_canvas.create_image(575,385, image=bgimage11)

    barid = StringVar()
    
    bar_id_entry = ctk.CTkEntry(add_stock_frame, height=40,width=330, fg_color='#f2f3f4', text_color='black',
                                border_color='#4B54F8', bg_color='white', corner_radius=20, 
                                font=('Poppins', 22), textvariable= barid)
    bar_id_entry.place(x = 700, y = 125)
    
    productname = StringVar()
    
    productname_entry = ctk.CTkEntry(add_stock_frame, height=40,width=290, fg_color='#f2f3f4', text_color='black',
                                border_color='#4B54F8', bg_color='white', corner_radius=20, 
                                font=('Poppins', 22), textvariable= productname)
    productname_entry.place(x = 730, y = 195)
    
    productqty = StringVar()
    
    productQuantity_entry = ctk.CTkEntry(add_stock_frame, height=40,width=290, fg_color='#f2f3f4', text_color='black',
                                border_color='#4B54F8', bg_color='white', corner_radius=20, 
                                font=('Poppins', 22), textvariable= productqty)
    productQuantity_entry.place(x = 730, y = 260)
    
    cp = StringVar()
    
    cp_entry = ctk.CTkEntry(add_stock_frame, height=40,width=290, fg_color='#f2f3f4', text_color='black',
                                border_color='#4B54F8', bg_color='white', corner_radius=20, 
                                font=('Poppins', 22),  textvariable= cp)
    cp_entry.place(x = 730, y = 325)
    
    sp = StringVar()
        
    sp_entry = ctk.CTkEntry(add_stock_frame, height=40,width=290, fg_color='#f2f3f4', text_color='black',
                                border_color='#4B54F8', bg_color='white', corner_radius=20, 
                                font=('Poppins', 22), textvariable = sp)
    sp_entry.place(x = 730, y = 390)
    
    tax = StringVar()
    
    tax_entry = ctk.CTkEntry(add_stock_frame, height=40,width=290, fg_color='#f2f3f4', text_color='black',
                                border_color='#4B54F8', bg_color='white', corner_radius=20, 
                                font=('Poppins', 22), textvariable= tax)
    tax_entry.place(x = 730, y = 455)
    
    mfdate = StringVar()

    mf_date_entry = ctk.CTkEntry(add_stock_frame, height=40,width=200, fg_color='#f2f3f4', text_color='black',
                                border_color='#4B54F8', bg_color='white', corner_radius=20, font=('Poppins', 22),
                                textvariable = mfdate)
    mf_date_entry.place(x = 780, y = 530)
    
    calendar_pil_image = Image.open(cwd+"\\uiux\\Calendar.png")
    calendar_icon = ctk.CTkImage(light_image=calendar_pil_image, dark_image=calendar_pil_image, size=(32, 32))
    
    calendar_button = ctk.CTkButton(add_stock_frame, image=calendar_icon, text="", width=40, height=40, fg_color= "white",
                                    bg_color="white", hover_color="white", cursor = 'hand2', command=open_calendar)
    calendar_button.place( x = 980, y = 530)
    
    expdate = StringVar()
    
    exp_date_entry = ctk.CTkEntry(add_stock_frame, height=40,width=200, fg_color='#f2f3f4', text_color='black',
                                border_color='#4B54F8', bg_color='white', corner_radius=20, font=('Poppins', 22),
                                textvariable= expdate)
    exp_date_entry.place(x = 780, y = 600)

    calendar_pil_image1 = Image.open(cwd+"\\uiux\\Calendar.png")
    calendar_icon1 = ctk.CTkImage(light_image=calendar_pil_image1, dark_image=calendar_pil_image1, size=(32, 32))
    
    calendar_button = ctk.CTkButton(add_stock_frame, image=calendar_icon1, text="", width=40, height=40, fg_color= "white",
                                    bg_color="white", hover_color="white", cursor = 'hand2', command=open_calendar1)
    calendar_button.place( x = 980, y = 600)
    
    
    add_stocks_button = ctk.CTkButton(add_stock_frame, text="Add", font=('arial', 26, 'bold'), height=45, width = 40, cursor='hand2',
                                fg_color='#2ecc71', bg_color='white', text_color='white', hover_color='black', 
                                corner_radius=10, command= add_stock)
    add_stocks_button.place(x = 595, y = 660)
    back_button = ctk.CTkButton(add_stock_frame, text="Back", font=('arial', 26, 'bold'), height=45, width = 40, cursor='hand2',
                                fg_color='#8e44ad', bg_color='white', text_color='white', 
                                hover_color='black', corner_radius=10, command=back_stock_entry)
    back_button.place(x = 750, y = 660)
    clear_button = ctk.CTkButton(add_stock_frame, text="Clear", font=('arial', 26, 'bold'), height=45, width = 40, cursor='hand2',
                                fg_color='#e74c3c', bg_color='white', text_color='white', 
                                hover_color='black', corner_radius=10, command= clear_inventory_entries
                                )
    clear_button.place(x = 900, y = 660)
    
    # =======================<<<<<<<<<<<< Add Stock Display INTERFACE FROM HERE >>>>>>>>>>>>>>>=============================

    update_stock_frame=Frame(dashboard_frame,height=770,width=1150,bg='blue')
    update_stock_frame.propagate(False)
    # inventory_display.pack(side='left')
    
    update_stock_canvas=Canvas(update_stock_frame,height=770,width=1150,bg='blue',bd=0,highlightthickness=0, relief='ridge')
    update_stock_canvas.propagate(False)
    update_stock_canvas.pack()


    imagepath12=cwd+"\\uiux\\update_stock3.png"
    openphoto12=Image.open(imagepath12).resize((1150,770))
    bgimage12=ImageTk.PhotoImage(openphoto12)
    update_stock_canvas.create_image(575,385, image=bgimage12)
    
    item_id_in_update = StringVar()
    
    bar_entry = ctk.CTkEntry(update_stock_frame, height=40,width=303, fg_color='white', text_color='black',
                             bg_color='white', corner_radius=20, font=('Poppins', 22), border_color='white',
                             textvariable= item_id_in_update)
    bar_entry.place(x = 585, y = 118)
    
    find_result_label = Label(update_stock_frame, text='', font=('poppins', 18, 'bold'), bg ='#4B54F8',
                              anchor='w', fg='white')
    
    
    id_src_button = ctk.CTkButton(update_stock_frame, text="Search", font=('Popins', 26, 'bold'), height=45, width = 40, cursor='hand2',
                                fg_color='#4B54F8', bg_color='white', text_color='white', 
                                hover_color='black', corner_radius=10, command= find_item_in_update)
    id_src_button.place(x = 940, y = 118)
    
    
    found_product_label = ctk.CTkLabel(update_stock_frame, text="", font=("poppins", 18), 
                                       fg_color = 'white', text_color='white')
    found_product_label.place( x = 585, y = 190)
    
    newproductval = StringVar()
    
    new_product_entry = ctk.CTkEntry(update_stock_frame, height=40,width=290, fg_color='#f2f3f4', text_color='black',
                                border_color='#4B54F8', bg_color='white', corner_radius=20, 
                                font=('Poppins', 22), textvariable=newproductval)
    new_product_entry.place(x = 800, y = 230)
    
    newcpval = StringVar()
    
    new_cp_entry = ctk.CTkEntry(update_stock_frame, height=40,width=290, fg_color='#f2f3f4', text_color='black',
                                border_color='#4B54F8', bg_color='white', corner_radius=20, 
                                font=('Poppins', 22), textvariable= newcpval)
    new_cp_entry.place(x = 800, y = 285)
    
    newspval = StringVar()
    
    new_sp_entry = ctk.CTkEntry(update_stock_frame, height=40,width=290, fg_color='#f2f3f4', text_color='black',
                                border_color='#4B54F8', bg_color='white', corner_radius=20, 
                                font=('Poppins', 22), textvariable=newspval)
    new_sp_entry.place(x = 800, y = 345)
    
    newtaxval = StringVar()
    
    new_tax_entry = ctk.CTkEntry(update_stock_frame, height=40,width=290, fg_color='#f2f3f4', text_color='black',
                                border_color='#4B54F8', bg_color='white', corner_radius=20, 
                                font=('Poppins', 22), textvariable=newtaxval)
    new_tax_entry.place(x = 800, y = 405)
    
    newdiscountval = StringVar()
    
    new_discount_entry = ctk.CTkEntry(update_stock_frame, height=40,width=290, fg_color='#f2f3f4', text_color='black',
                                border_color='#4B54F8', bg_color='white', corner_radius=20, 
                                font=('Poppins', 22), textvariable = newdiscountval)
    new_discount_entry.place(x = 800, y = 465)
    
    
    update_button = ctk.CTkButton(update_stock_frame, text="Update", font=('Popins', 26, 'bold'), height=45, width = 40, cursor='hand2',
                                fg_color='#4B54F8', bg_color='white', text_color='white', 
                                hover_color='black', corner_radius=10, command=update_product)
    update_button.place(x = 615, y = 530)
    
    
    update_back_button = ctk.CTkButton(update_stock_frame, text="Back", font=('Popins', 26, 'bold'), height=45, width = 40, cursor='hand2',
                                fg_color='#4B54F8', bg_color='white', text_color='white', 
                                hover_color='black', corner_radius=10, command=back_update_entry)
    update_back_button.place(x = 790, y = 530)
    
    update_clear_button = ctk.CTkButton(update_stock_frame, text="Clear", font=('Popins', 26, 'bold'), height=45, width = 40, cursor='hand2',
                                fg_color='#4B54F8', bg_color='white', text_color='white', 
                                hover_color='black', corner_radius=10, command = delete_update_entries)
    update_clear_button.place(x = 935, y = 530)
    
    # =======================<<<<<<<<<<<< Add stock analytics INTERFACE FROM HERE >>>>>>>>>>>>>>>=============================

    stock_analytics_frame=Frame(dashboard_frame,height=770,width=1150,bg='blue')
    stock_analytics_frame.propagate(False)
    # inventory_display.pack(side='left')
    
    stock_analytics_canvas=Canvas(stock_analytics_frame,height=770,width=1150,bg='blue',bd=0,highlightthickness=0, relief='ridge')
    stock_analytics_canvas.propagate(False)
    stock_analytics_canvas.pack()


    imagepath13=cwd+"\\uiux\\analytics.png"
    openphoto13=Image.open(imagepath13).resize((1150,770))
    bgimage13=ImageTk.PhotoImage(openphoto13)
    stock_analytics_canvas.create_image(575,385, image=bgimage13)
    
    
    analytics_board = Text(stock_analytics_frame, wrap="none", height=25, width=136, relief='ridge', bd=3,
                           state= DISABLED, fg='black')
    analytics_board.place(x = 20, y = 42)
    
    v_scroll = ctk.CTkScrollbar(stock_analytics_frame, orientation="vertical", command=analytics_board.yview, bg_color='white', 
                                height=405)
    v_scroll.place(x = 1120, y = 42)
    
    h_scroll = ctk.CTkScrollbar(stock_analytics_frame, orientation="horizontal", command=analytics_board.xview,
                                width=1090, bg_color='white')
    h_scroll.place(x = 24, y = 455)
    
    analytics_board.config(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)
    
    analytics_src_val = StringVar()
    
    analytics_src_entry = ctk.CTkEntry(stock_analytics_frame, font=('Poppins', 18), fg_color = 'white',
                                       bg_color='white', border_color='#4B54F8', corner_radius=20,
                                       height=30, width=250, textvariable=analytics_src_val)
    analytics_src_entry.place(x = 350, y = 520)
    
    analytics_src_button = ctk.CTkButton(stock_analytics_frame, text="Search", font=('Popins', 18, 'bold'), height=15, width = 175, cursor='hand2',
                                fg_color='#17a589', bg_color='white', text_color='white', 
                                hover_color='black', corner_radius=30, anchor='center', command = analytics_idsrc_display)
    analytics_src_button.place( x = 630, y = 522)
    
    analytics_option = ['All', 'Cost, Selling Price', 'Product, Selling Price','Product, Cost Price', 'Product, Discount', 'Product, Tax', 'Product, Tax, Discount']
    
    analytics_option_menu = ctk.CTkOptionMenu(stock_analytics_frame, values=analytics_option, bg_color='white', fg_color='#4B54F8',
                                              font=('Poppins', 15, 'bold'), text_color='white', width= 250,
                                              button_color='black', button_hover_color='grey', command=option_selected)
    analytics_option_menu.set(analytics_option[0])  # Set default option
    analytics_option_menu.place(x = 450, y = 570)

    analytics_back_button = ctk.CTkButton(stock_analytics_frame, text="Back", font=('Popins', 18, 'bold'), height=15, width = 120, cursor='hand2',
                                fg_color='#e74c3c', bg_color='white', text_color='white', 
                                hover_color='black', corner_radius=30, anchor='center', command = analytics_back)
    analytics_back_button.place( x = 315, y = 640)
    
    analytics_reset_button = ctk.CTkButton(stock_analytics_frame, text="Reset", font=('Popins', 18, 'bold'), height=15, width = 120, cursor='hand2',
                                fg_color='#2e86c1', bg_color='white', text_color='white', 
                                hover_color='black', corner_radius=30, anchor='center', command = fetch_and_display_inventory)
    analytics_reset_button.place( x = 450, y = 640)
    
    analytics_print_button = ctk.CTkButton(stock_analytics_frame, text="Print", font=('Popins', 18, 'bold'), height=15, width = 120, cursor='hand2',
                                fg_color='#7d3c98', bg_color='white', text_color='white', 
                                hover_color='black', corner_radius=30, anchor='center')
    analytics_print_button.place( x = 585, y = 640)
    
    analytics_save_button = ctk.CTkButton(stock_analytics_frame, text="Save", font=('Popins', 18, 'bold'), height=15, width = 120, cursor='hand2',
                                fg_color='#616a6b', bg_color='white', text_color='white', 
                                hover_color='black', corner_radius=30, anchor='center', command = save_analytics_display)
    analytics_save_button.place( x = 720, y = 640)
    # =======================<<<<<<<<<<<< alert Display INTERFACE FROM HERE >>>>>>>>>>>>>>>=============================

    alert_display=Frame(dashboard_frame,height=770,width=1150,bg='green')
    alert_display.propagate(False)
    # alert_display.pack(side='left')
    alert_canvas=Canvas(alert_display,height=770,width=1150,bg='blue',bd=0,highlightthickness=0, relief='ridge')
    alert_canvas.propagate(False)
    alert_canvas.pack()


    imagepath15=cwd+"\\uiux\\alert11.png"
    openphoto15=Image.open(imagepath15).resize((1150,770))
    bgimage15=ImageTk.PhotoImage(openphoto15)
    alert_canvas.create_image(575,385, image=bgimage15)
    
    expiring_frame = Frame(alert_canvas, bg='red', height=500, width=800)
    expiring_frame.propagate(False)
    expiring_frame.pack(side='left', anchor='nw', padx=20, pady=(170,0))
    
    
    columns = ("SLNO", "Product_ID", "Product_Name", "Expiring In")
    alert_tree_expire = ttk.Treeview(expiring_frame, columns=columns, show="headings")
    for col in columns:
        alert_tree_expire.heading(col, text=col)
    
    alert_tree_expire.column("SLNO", width=60, anchor="center")
    alert_tree_expire.column("Product_ID", width=100, anchor="center")
    alert_tree_expire.column("Product_Name", width=180, anchor="w")
    alert_tree_expire.column("Expiring In", width=100, anchor="center")
    # tree.column("Stock left", width=100, anchor="center")
    
    
    alert_tree_expire.bind("<<TreeviewSelect>>", on_row_selected_expire)
    
    # Add vertical scrollbar
    v_scrollbar = ttk.Scrollbar(expiring_frame, orient="vertical", command=alert_tree_expire.yview)
    alert_tree_expire.configure(yscrollcommand=v_scrollbar.set)

    # Place Treeview and Scrollbar side-by-side
    alert_tree_expire.pack(side="left", fill="both", expand=True)
    v_scrollbar.pack(side="right", fill="y")
    
    alert_info_frame = Frame(alert_canvas, height=500, width=270)
    alert_info_frame.propagate(False)
    alert_info_frame.pack(side='right', anchor='ne', padx=(0,20), pady=(170, 0))
    
    alert_textbox = ctk.CTkTextbox(alert_info_frame, font=('Poppins', 12), border_color='#4B54F8', border_width=2, scrollbar_button_color='#4B54F8',
                                   scrollbar_button_hover_color='grey', fg_color="#F5F6FF")
    alert_textbox.pack(fill = 'both', expand = True)
    
    textbox_clear_button = ctk.CTkButton(alert_canvas, text="Clear", font=('Poppins', 16, 'bold'), text_color='black',
                            fg_color='#4B54F8', cursor='hand2', bg_color='white', border_color='#4B54F8', hover_color='grey',
                            command=clear_alert_texbox)
    textbox_clear_button.place(x = 920, y = 690)
    
    exchange_label_button = ctk.CTkButton(alert_canvas, text="Show Stock Alert", font=('Poppins', 16, 'bold'), text_color='black',
                            fg_color='white', cursor='hand2', bg_color='white', border_color='#4B54F8', hover_color='grey',
                            command= switch_to_stock)
    exchange_label_button.place(x = 630, y = 690)
    
    exchange_pil_image = Image.open(cwd+"\\uiux\\exchange.png")
    exchange_icon = ctk.CTkImage(light_image=exchange_pil_image, dark_image=exchange_pil_image, size=(30, 30))
    
    exchange_button = ctk.CTkButton(alert_canvas, image=exchange_icon, text="", width=40, height=40, fg_color= "white",
                                    bg_color="white", hover_color="white", cursor = 'hand2', command= switch_to_stock)
    exchange_button.place( x = 770, y = 682)
    
    stock_alert_frame = Frame(alert_canvas, bg='red', height=500, width=800)
    stock_alert_frame.propagate(False)
    
    columns2 = ("SLNO", "Product_ID", "Product_Name", "Stock left")
    alert_tree_stock = ttk.Treeview(stock_alert_frame, columns=columns2, show="headings")
    for col in columns2:
        alert_tree_stock.heading(col, text=col)
    
    alert_tree_stock.column("SLNO", width=60, anchor="center")
    alert_tree_stock.column("Product_ID", width=100, anchor="center")
    alert_tree_stock.column("Product_Name", width=180, anchor="w")
    # alert_tree_stock.column("Expiring In", width=100, anchor="center")
    alert_tree_stock.column("Stock left", width=100, anchor="center")
    
    
    alert_tree_stock.bind("<<TreeviewSelect>>", on_row_selected_stock)
    
    # Add vertical scrollbar
    v_scrollbar = ttk.Scrollbar(stock_alert_frame, orient="vertical", command=alert_tree_stock.yview)
    alert_tree_stock.configure(yscrollcommand=v_scrollbar.set)

    # Place Treeview and Scrollbar side-by-side
    alert_tree_stock.pack(side="left", fill="both", expand=True)
    v_scrollbar.pack(side="right", fill="y")
    
    # =======================<<<<<<<<<<<< billing Display INTERFACE FROM HERE >>>>>>>>>>>>>>>=============================
    
    billing_display=Frame(dashboard_frame,height=770,width=1150,bg='yellow')
    billing_display.propagate(False)
    # billing_display.pack(side='left')
    
    billing_canvas=Canvas(billing_display,height=770,width=1150,bg='blue',bd=0,highlightthickness=0, relief='ridge')
    billing_canvas.propagate(False)
    billing_canvas.pack()


    imagepath16=cwd+"\\uiux\\invoice8.png"
    openphoto16=Image.open(imagepath16).resize((1150,770))
    bgimage16=ImageTk.PhotoImage(openphoto16)
    billing_canvas.create_image(575,385, image=bgimage16)
    
    barcodevalue = StringVar()
    
    barcode_id_entry = ctk.CTkEntry(billing_display, width = 313, height= 35, fg_color ='white', font=('poppins', 18), corner_radius = 13,
                                    border_color='white', bg_color='white', textvariable=barcodevalue)
    barcode_id_entry.place(x = 28, y = 199)
    

    billing_columns = (
        "SLNO", "Product ID", "Product Name", "Quantity", "SP",
        "SGST", "SGST(Rs)", "CGST", "CGST(Rs)", "Discount(%)", "Amount(GST Inc)"
    )

    # Create Treeview
    billing_tree = ttk.Treeview(billing_display, columns=billing_columns, show="headings", height=20)

    # Define column headings
    for bill_col in billing_columns:
        if bill_col == 'SLNO':
            billing_tree.heading(bill_col, text=bill_col)
            billing_tree.column(bill_col, anchor="center", width=50)
        else:
            billing_tree.heading(bill_col, text=bill_col)
            billing_tree.column(bill_col, anchor="center", width=100)

    # # Example data (you can remove this and add your own later)
    

    # Add vertical scrollbar
    billing_scrollbar = ttk.Scrollbar(billing_display, orient="vertical", command=billing_tree.yview)
    billing_tree.configure(yscrollcommand=billing_scrollbar.set)
    billing_scrollbar.place(x = 1080, y = 310, height= 430)

    # Pack the treeview
    billing_tree.place(x = 25, y = 310)

    barcode_id_entry.bind("<Return>", billing_tree_insert)
    
    bill_find_entry = ctk.CTkEntry(billing_display, width = 200, height= 25, fg_color ='white', font=('poppins', 18), corner_radius = 13,
                                    border_color='black', bg_color='white')
    bill_find_entry.place(x = 810, y = 55)
    
    bill_find_button = ctk.CTkButton(billing_display, font=('Poppins', 18, 'bold'), text="Find Bill", height=25, width = 100, fg_color="#4B54F8",
                                     bg_color="white", corner_radius=13, hover_color='black', text_color="white")
    bill_find_button.place(x = 860, y = 87)
    
    buyer_name_entry = ctk.CTkEntry(billing_display, width =220, height= 25, fg_color ='white', font=('poppins', 18), corner_radius = 13,
                                    border_color='black', bg_color='white')
    buyer_name_entry.place(x = 570, y = 55)
    
    buyer_phone_entry = ctk.CTkEntry(billing_display, width =190, height= 25, fg_color ='white', font=('poppins', 18), corner_radius = 13,
                                    border_color='black', bg_color='white')
    buyer_phone_entry.place(x = 600, y = 87)
    
    buyer_address_entry = ctk.CTkEntry(billing_display, width =200, height= 25, fg_color ='white', font=('poppins', 18), corner_radius = 13,
                                    border_color='black', bg_color='white')
    buyer_address_entry.place(x = 590, y = 118)
    
    cash_var = IntVar(value=1)
    
    cash_checkbox = ctk.CTkCheckBox(billing_display, text="Cash", text_color="black", hover_color="black",
                                   checkmark_color="white", bg_color="White", checkbox_height=25, checkbox_width=25,
                                   font=('poppins', 16), border_color="grey", variable= cash_var,
                                   command= on_cash_click)
    cash_checkbox.place(x = 820, y = 122)
    
    upi_var = IntVar()
    
    upi_checkbox = ctk.CTkCheckBox(billing_display, text="UPI", text_color="black", hover_color="black",
                                   checkmark_color="white", bg_color="White", checkbox_height=25, checkbox_width=25,
                                   font=('poppins', 16), border_color="grey",variable=upi_var,
                                   command= on_upi_click)
    upi_checkbox.place(x = 940, y = 122)
    
    reset_info = ctk.CTkButton(billing_display, font=('Poppins', 18, 'bold'), text="Reset Information", height=35, width = 200, fg_color="#F8514B",
                                     bg_color="white", corner_radius=13, hover_color='black', text_color="white",
                                     command = clear_bill_treeview)
    reset_info.place(x = 610, y = 158)
    
    generate_bill_btn = ctk.CTkButton(billing_display, font=('Poppins', 18, 'bold'), text="Generate Bill", height=35, width = 200, fg_color="#28F13F",
                                     bg_color="white", corner_radius=13, hover_color='black', text_color="white",
                                     command = gen_bill_win)
    generate_bill_btn.place(x = 610, y = 198)
    
    def_var = IntVar(value=1)
    def_checkbox = ctk.CTkCheckBox(billing_display, text="Default", text_color="black", hover_color="black",
                                   checkmark_color="white", bg_color="White", checkbox_height=25, checkbox_width=25,
                                   font=('poppins', 16), border_color="grey", variable= def_var, command= on_checkbox1_click)
    def_checkbox.place(x = 820, y = 200)
    
    a4_var = IntVar()
    
    a4_checkbox = ctk.CTkCheckBox(billing_display, text="A4 size", text_color="black", hover_color="black",
                                   checkmark_color="white", bg_color="White", checkbox_height=25, checkbox_width=25,
                                   font=('poppins', 16), border_color="grey",variable=a4_var, command=on_checkbox2_click)
    a4_checkbox.place(x = 940, y = 200)
    
    

    
    # =======================<<<<<<<<<<<< supply Display INTERFACE FROM HERE >>>>>>>>>>>>>>>=============================
    
    supply_display=Frame(dashboard_frame,height=770,width=1150,bg='pink')
    supply_display.propagate(False)
    # supply_display.pack(side='left')
    supply_canvas=Canvas(supply_display,height=770,width=1150,bg='blue',bd=0,highlightthickness=0, relief='ridge')
    supply_canvas.propagate(False)
    supply_canvas.pack()


    imagepath18=cwd+"\\uiux\\supp_undev.png"
    openphoto18=Image.open(imagepath18).resize((1150,770))
    bgimage18=ImageTk.PhotoImage(openphoto18)
    supply_canvas.create_image(575,385, image=bgimage18)
    
    # =======================<<<<<<<<<<<< history Display INTERFACE FROM HERE >>>>>>>>>>>>>>>=============================
    
    history_display=Frame(dashboard_frame,height=770,width=1150,bg='violet')
    history_display.propagate(False)
    # history_display.pack(side='left')
    
    history_canvas=Canvas(history_display,height=770,width=1150,bg='blue',bd=0,highlightthickness=0, relief='ridge')
    history_canvas.propagate(False)
    history_canvas.pack()


    imagepath17=cwd+"\\uiux\\history.png"
    openphoto17=Image.open(imagepath17).resize((1150,770))
    bgimage17=ImageTk.PhotoImage(openphoto17)
    history_canvas.create_image(575,385, image=bgimage17)
    
    
    local_history_btn = ctk.CTkButton(history_display, text="Local", font=('poppins', 16, 'bold'), fg_color="#3498db", text_color="white",
                                      bg_color="white", width=117, height=35, cursor = "hand2", hover_color="Black",
                                      command=show_local_history)
    local_history_btn.place(x = 450, y = 213)
    
    
    global_history_btn = ctk.CTkButton(history_display, text="Global", font=('poppins', 16, 'bold'), fg_color="#8e44ad", text_color="white",
                                      bg_color="white", width=117, height=35, cursor = "hand2", hover_color="Black",
                                      command=show_global_history)
    global_history_btn.place(x = 585, y = 213)
    
    local_history_frame = Frame(history_canvas, bg="red", relief='ridge', bd=3, height=500)
    local_history_frame.propagate(False)
    local_history_frame.pack(side='bottom', anchor='s', fill=X, expand=True, padx=10, pady=(0, 10))
    
    
    lc_history_columns = ("Product_ID", "Product_Name", "Date", "Time", "Amount(Rs)", "Action")

    # Create the Treeview
    history_tree = ttk.Treeview(local_history_frame, columns=lc_history_columns, show="headings")

    # Define headings
    for cols in lc_history_columns:
        if cols == "Product_Name":
            history_tree.heading(cols, text=cols)
            history_tree.column(cols, width=150, anchor="center")
        elif cols == "Time":
            history_tree.heading(cols, text=cols)
            history_tree.column(cols, width=50, anchor="center")
        else:
            history_tree.heading(cols, text=cols)
            history_tree.column(cols, width=100, anchor="center")

    vsb = ttk.Scrollbar(local_history_frame, orient="vertical", command=history_tree.yview)
    history_tree.configure(yscrollcommand=vsb.set)

    # Use pack to position the Treeview and Scrollbar
    vsb.pack(side="right", fill="y")
    history_tree.pack(side="left", expand=True, fill="both")
    
    global_history_frame = Frame(history_canvas, bg="blue", relief='ridge', bd=3, height=500)
    global_history_frame.propagate(False)
    # global_history_frame.pack(side='bottom', anchor='s', fill=X, expand=True, padx=10, pady=(0, 10))

    glo_history_columns = ("Product_ID", "Email", "Product_Name", "Date", "Time", "Amount(Rs)", "Action")

    # Create the Treeview
    glo_history_tree = ttk.Treeview(global_history_frame, columns=glo_history_columns, show="headings")

    # Define headings
    for cols in glo_history_columns:
        if cols == "Product_Name":
            glo_history_tree.heading(cols, text=cols)
            glo_history_tree.column(cols, width=150, anchor="center")
        elif cols == "Time":
            glo_history_tree.heading(cols, text=cols)
            glo_history_tree.column(cols, width=50, anchor="center")
        else:
            glo_history_tree.heading(cols, text=cols)
            glo_history_tree.column(cols, width=100, anchor="center")

    vsb2 = ttk.Scrollbar(global_history_frame, orient="vertical", command=glo_history_tree.yview)
    glo_history_tree.configure(yscrollcommand=vsb2.set)

    # Use pack to position the Treeview and Scrollbar
    vsb2.pack(side="right", fill="y")
    glo_history_tree.pack(side="left", expand=True, fill="both")
    # =======================<<<<<<<<<<<< setting Display INTERFACE FROM HERE >>>>>>>>>>>>>>>=============================
    
    setting_display=Frame(dashboard_frame,height=770,width=1150,bg='brown')
    setting_display.propagate(False)
    # setting_display.pack(side='left')
    
    setting_canvas=Canvas(setting_display,height=770,width=1150,bg='blue',bd=0,highlightthickness=0, relief='ridge')
    setting_canvas.propagate(False)
    setting_canvas.pack()


    imagepath19=cwd+"\\uiux\\setting_undev.png"
    openphoto19=Image.open(imagepath19).resize((1150,770))
    bgimage19=ImageTk.PhotoImage(openphoto19)
    setting_canvas.create_image(575,385, image=bgimage19)
    
    # =======================<<<<<<<<<<<< cc Display INTERFACE FROM HERE >>>>>>>>>>>>>>>=============================
    
    cc_display=Frame(dashboard_frame,height=770,width=1150,bg='grey')
    cc_display.propagate(False)
    # cc_display.pack(side='left')
    cc_canvas=Canvas(cc_display,height=770,width=1150,bg='blue',bd=0,highlightthickness=0, relief='ridge')
    cc_canvas.propagate(False)
    cc_canvas.pack()


    imagepath14=cwd+"\\uiux\\cc3.png"
    openphoto14=Image.open(imagepath14).resize((1150,770))
    bgimage14=ImageTk.PhotoImage(openphoto14)
    cc_canvas.create_image(575,385, image=bgimage14)
    
    
    cc_textbox = ctk.CTkTextbox(cc_display, width=400, height=200, bg_color = 'white', fg_color='#d6eaf8', 
                                scrollbar_button_color='#4B54F8', border_color='red', font=('poppins', 12, 'bold'),border_width=1)
    cc_textbox.place(x = 640, y = 200)

    cc_send_button = ctk.CTkButton(cc_display, text="Send Message", font=('poppins', 18, 'bold'),command=get_cc_text, 
                                   fg_color='#4B54F8', bg_color='white', text_color='white', corner_radius=20,
                                   hover_color='black')
    cc_send_button.place(x = 775, y = 450)

    win.mainloop()
    
else:
    messagebox.showerror('Connection Error', 'Please Check your internet connection!')


